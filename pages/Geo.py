import streamlit as st
import pandas as pd
import google.generativeai as genai
from fpdf import FPDF
import openpyxl
import os
import base64
from datetime import datetime
import tempfile
import re
import json
import requests
import urllib.parse
import io
from PIL import Image, ImageDraw, ImageFont
from pathlib import Path
from typing import Optional

# --- Geodata Online client (if credentials configured) ---
try:
    from geodata_client import GeodataOnlineClient, geodata_buildings_to_neighbors
    _gdo = GeodataOnlineClient()
    _HAS_GDO = _gdo.is_available()
    _GDO_TOKEN_OK = False
    if _HAS_GDO:
        try:
            _gdo.get_token()
            _GDO_TOKEN_OK = True
        except Exception:
            _GDO_TOKEN_OK = False
except Exception:
    _gdo = None
    _HAS_GDO = False
    _GDO_TOKEN_OK = False

# --- Auth integration (for saving reports to user account) ---
try:
    import builtly_auth
    _HAS_AUTH = True
except ImportError:
    _HAS_AUTH = False

# Restore auth session on module pages (session_state may be lost during navigation)
if _HAS_AUTH:
    if not st.session_state.get("user_authenticated"):
        builtly_auth.try_restore_from_browser()
    elif st.session_state.get("_sb_access_token"):
        builtly_auth.restore_session()


# ── OPEN NORWEGIAN GEODATA APIs (no auth required) ─────────────────────────

# Norge i Bilder via Geodata Online proxy (uses same token as Geodata Online — no GeoID needed!)
# Discovered at services.geodataonline.no/arcgis/rest/services/Wms
# SecureWms endpoint format — try WmsServer (ArcGIS standard for SecureWms type)
NIB_PROSJEKTER_GDO_URLS = [
    "https://services.geodataonline.no/arcgis/services/Wms/Geonorge_NiB-prosjekter/WmsServer",
    "https://services.geodataonline.no/arcgis/services/Wms/Geonorge_NiB-prosjekter/WMSServer",
    "https://services.geodataonline.no/arcgis/services/Wms/Geonorge_NiB-prosjekter/SecureWmsServer",
]

def _wms_get_map(base_url: str, layers: str, bbox_25833: tuple,
                 width: int = 800, height: int = 800,
                 srs: str = "EPSG:25833", version: str = "1.1.1",
                 extra_params: dict = None,
                 auth: tuple = None, timeout: int = 12) -> Optional[Image.Image]:
    """Generic WMS GetMap helper. Returns PIL Image or None.
    
    Args:
        auth: Optional (username, password) tuple for HTTP Basic Auth
    """
    minx, miny, maxx, maxy = bbox_25833
    params = {
        "service": "WMS", "request": "GetMap", "version": version,
        "layers": layers, "styles": "",
        "srs" if version < "1.3" else "crs": srs,
        "bbox": f"{minx},{miny},{maxx},{maxy}",
        "width": str(width), "height": str(height),
        "format": "image/png", "transparent": "true",
    }
    if extra_params:
        params.update(extra_params)
    try:
        resp = requests.get(base_url, params=params, timeout=timeout, auth=auth)
        if resp.status_code == 200 and len(resp.content) > 2000:
            if b"ServiceException" in resp.content[:1000] or b"<html" in resp.content[:200].lower():
                return None
            return Image.open(io.BytesIO(resp.content)).convert("RGBA")
    except Exception:
        pass
    return None


def _geodata_online_map(service: str, bbox: tuple, width: int = 800, height: int = 800) -> Optional[Image.Image]:
    """Fetch a map image from Geodata Online MapServer (requires credentials)."""
    if not _HAS_GDO:
        return None
    try:
        img_bytes = _gdo._image_export(service, bbox, width=width, height=height)
        if img_bytes:
            return Image.open(io.BytesIO(img_bytes)).convert("RGB")
    except Exception:
        pass
    return None


def _coords_from_address(adresse: str, kommune: str, gnr: str, bnr: str) -> Optional[tuple]:
    """Resolve address/gnr-bnr to UTM33 coordinates via Geonorge."""
    # Method 1: Address search
    queries = []
    adr = (adresse or "").replace(",", "").strip()
    kom = (kommune or "").replace(",", "").strip()
    if adr and kom:
        queries.append(f"{adr} {kom}")
    if adr:
        queries.append(adr)
    if gnr and bnr and kom:
        queries.append(f"{kom} {gnr}/{bnr}")

    for q in queries:
        try:
            url = f"https://ws.geonorge.no/adresser/v1/sok?sok={urllib.parse.quote(q)}&fuzzy=true&utkoordsys=25833&treffPerSide=1"
            resp = requests.get(url, timeout=5)
            if resp.status_code == 200:
                hits = resp.json().get("adresser", [])
                if hits:
                    rp = hits[0].get("representasjonspunkt", {})
                    nord, ost = rp.get("nord"), rp.get("øst")
                    if nord and ost:
                        return (float(ost), float(nord))
        except Exception:
            pass

    # Method 2: WFS parcel lookup by gnr/bnr (same as Project Setup uses)
    if gnr and bnr and kom:
        coords = _coords_from_gnr_bnr(kom, gnr, bnr)
        if coords:
            return coords

    return None


def _resolve_kommunenummer(kommune: str) -> Optional[str]:
    """Resolve kommune name to kommunenummer."""
    s = kommune.strip()
    if s.isdigit() and len(s) >= 3:
        return s.zfill(4)
    try:
        resp = requests.get("https://ws.geonorge.no/kommuneinfo/v1/kommuner", timeout=5)
        if resp.status_code == 200:
            for k in resp.json():
                if k.get("kommunenavn", "").lower() == s.lower():
                    return k.get("kommunenummer")
    except Exception:
        pass
    return None


def _coords_from_gnr_bnr(kommune: str, gnr: str, bnr: str) -> Optional[tuple]:
    """Get centroid coordinates from Kartverket WFS using gnr/bnr."""
    knr = _resolve_kommunenummer(kommune)
    if not knr:
        return None

    gnr_clean = str(gnr).strip()
    bnr_clean = str(bnr).strip()
    cql = f"kommunenummer='{knr}' AND gardsnummer={gnr_clean} AND bruksnummer={bnr_clean}"

    for wfs_url, layer in [
        ("https://wfs.geonorge.no/skwms1/wfs.matrikkelen-teig", "matrikkelen-teig:Teig"),
        ("https://wfs.geonorge.no/skwms1/wfs.matrikkelkart", "matrikkelkart:Teig"),
    ]:
        try:
            resp = requests.get(wfs_url, params={
                "service": "WFS", "version": "2.0.0", "request": "GetFeature",
                "typenames": layer, "srsName": "EPSG:25833",
                "outputFormat": "application/json", "cql_filter": cql,
            }, timeout=12)
            if resp.status_code != 200:
                resp = requests.get(wfs_url, params={
                    "service": "WFS", "version": "2.0.0", "request": "GetFeature",
                    "typenames": layer, "srsName": "EPSG:25833",
                    "outputFormat": "json", "cql_filter": cql,
                }, timeout=12)
            if resp.status_code == 200:
                features = resp.json().get("features", [])
                if features:
                    geom = features[0].get("geometry", {})
                    coords = geom.get("coordinates", [])
                    if coords:
                        # Calculate centroid from polygon
                        ring = coords[0] if isinstance(coords[0], list) and isinstance(coords[0][0], list) else coords
                        if isinstance(ring[0], list):
                            ring = ring[0]
                        xs = [p[0] for p in ring if isinstance(p, (list, tuple)) and len(p) >= 2]
                        ys = [p[1] for p in ring if isinstance(p, (list, tuple)) and len(p) >= 2]
                        if xs and ys:
                            return (sum(xs) / len(xs), sum(ys) / len(ys))
        except Exception:
            pass
    return None


def _bbox_from_center(x: float, y: float, radius_m: float = 250.0) -> tuple:
    return (x - radius_m, y - radius_m, x + radius_m, y + radius_m)


def fetch_ngu_losmasser(bbox: tuple) -> Optional[Image.Image]:
    """NGU løsmassekart — kvartærgeologisk kart."""
    for layers in ["Losmasse_flate", "0", "losmasser"]:
        img = _wms_get_map("https://geo.ngu.no/mapserver/LosijordsmasserWMS", layers, bbox)
        if img:
            return img
    return None

def fetch_ngu_berggrunn(bbox: tuple) -> Optional[Image.Image]:
    """NGU bergrunnskart."""
    for layers in ["Berggrunn_flate", "0", "berggrunn"]:
        img = _wms_get_map("https://geo.ngu.no/mapserver/BerggrunnWMS", layers, bbox)
        if img:
            return img
    return None

def fetch_ngu_radon(bbox: tuple) -> Optional[Image.Image]:
    """NGU radonkart — aktsomhetskart for radon."""
    for layers in ["Radonaktsomhet", "0", "radon"]:
        img = _wms_get_map("https://geo.ngu.no/mapserver/RadonWMS", layers, bbox)
        if img:
            return img
    return None

def fetch_nve_flom(bbox: tuple) -> Optional[Image.Image]:
    """NVE flomsonekart."""
    for layers in ["Flomsone", "0", "0,1,2"]:
        img = _wms_get_map("https://gis3.nve.no/map/services/Flomsoner1/MapServer/WMSServer", layers, bbox)
        if img:
            return img
    return None

def fetch_nve_kvikkleire(bbox: tuple) -> Optional[Image.Image]:
    """NVE kvikkleirekart."""
    img = _wms_get_map("https://gis3.nve.no/map/services/Kvikkleire2/MapServer/WMSServer", "0", bbox)
    return img

def fetch_nve_skred(bbox: tuple) -> Optional[Image.Image]:
    """NVE skredkart (alle typer)."""
    img = _wms_get_map("https://gis3.nve.no/map/services/SkredAktsomhet/MapServer/WMSServer", "0,1,2,3", bbox)
    return img

def fetch_kartverket_ortofoto(bbox: tuple) -> Optional[Image.Image]:
    """Kartverket ortofoto (Norge i Bilder)."""
    img = _wms_get_map("https://wms.geonorge.no/skwms1/wms.nib", "ortofoto", bbox)
    if img:
        return img.convert("RGB")
    return None


# ── HISTORISKE FLYBILDER ─────────────────────────────────────────────────────

def fetch_nib_historisk_ortofoto(bbox: tuple) -> tuple:
    """Fetch historical orthophotos from Norge i Bilder prosjekter WMS.
    
    Uses Geodata Online WMS proxy (services.geodataonline.no) which authenticates
    with the same Geodata Online token — no separate GeoID credentials needed.
    
    Strategy:
    1. GetCapabilities via Geodata Online proxy to discover available layers
    2. Parse for layers with year in name, pick oldest covering area
    3. Fetch GetMap for that layer
    
    Returns (PIL.Image, source_label, year_str) or (None, error_msg, None).
    """
    if not _GDO_TOKEN_OK:
        return None, "NIB prosjekter: Geodata Online ikke tilkoblet", None

    # Get Geodata Online token for SecureWms access
    try:
        token = _gdo.get_token()
    except Exception as e:
        return None, f"NIB prosjekter: token-feil — {str(e)[:60]}", None

    token_params = {"token": token}
    
    # Step 1: Discover available layers via GetCapabilities (try multiple URL variants)
    caps_resp = None
    working_url = None
    caps_params = {
        "service": "WMS",
        "request": "GetCapabilities",
        "version": "1.1.1",
        "token": token,
    }
    
    for url_candidate in NIB_PROSJEKTER_GDO_URLS:
        try:
            resp = requests.get(url_candidate, params=caps_params, timeout=30)
            if resp.status_code == 200 and len(resp.content) > 500:
                caps_resp = resp
                working_url = url_candidate
                break
            elif resp.status_code in (401, 403):
                return None, f"NIB prosjekter: ingen tilgang ({resp.status_code}) — kontakt Geodata for å aktivere NiB-prosjekter", None
        except requests.exceptions.Timeout:
            continue
        except Exception:
            continue

    if caps_resp is None:
        return None, f"NIB prosjekter: GetCapabilities feilet på alle URL-varianter (prøvde {len(NIB_PROSJEKTER_GDO_URLS)})", None

    # Step 2: Quick-parse layer names from XML
    # Full XML parsing is too slow for massive capabilities doc — use regex
    layer_candidates = []
    try:
        # WMS Name tags (with optional namespace prefix like <wms:Name>)
        name_pattern = re.compile(r'<(?:\w+:)?Name>\s*([^<]*(?:19|20)\d{2}[^<]*?)\s*</(?:\w+:)?Name>')
        for match in name_pattern.finditer(caps_resp.text):
            layer_name = match.group(1).strip()
            year_match = re.search(r'((?:19|20)\d{2})', layer_name)
            if year_match:
                year = int(year_match.group(1))
                if 1940 <= year <= 2010:
                    layer_candidates.append((year, layer_name))
    except Exception:
        pass

    if not layer_candidates:
        return None, f"NIB prosjekter: ingen historiske lag funnet (GetCapabilities {len(caps_resp.content)} bytes)", None

    # Sort by year (oldest first), deduplicate
    layer_candidates = list(dict.fromkeys(layer_candidates))
    layer_candidates.sort(key=lambda x: x[0])
    
    # Step 3: Try the oldest layers first (up to 8 attempts)
    for year, layer_name in layer_candidates[:8]:
        img = _wms_get_map(working_url, layer_name, bbox,
                           width=800, height=800, extra_params=token_params, timeout=15)
        if img:
            return img.convert("RGB"), f"Norge i Bilder ({layer_name})", str(year)

    tried = ", ".join(f"{name} ({yr})" for yr, name in layer_candidates[:3])
    return None, f"NIB prosjekter: {len(layer_candidates)} lag funnet men ingen ga bilde (prøvde: {tried})", None


def fetch_historical_ortofoto(bbox: tuple) -> dict:
    """Fetch historical aerial imagery from all available sources.
    
    Priority:
    1. Norge i Bilder prosjekter WMS (ekte historiske flybilder, 1947+)
    2. Geodata Online GeomapBilder variants (fallback — nyere årganger)
    
    Returns dict with keys: image, source, year, log
    """
    result = {"image": None, "source": "", "year": None, "log": []}

    # Expand bbox for historical imagery (parcel bounds might be too tight)
    minx, miny, maxx, maxy = bbox
    buf = 100.0  # 100m buffer
    expanded_bbox = (minx - buf, miny - buf, maxx + buf, maxy + buf)

    # Source 1: Norge i Bilder prosjekter — ekte historiske flybilder
    try:
        img, src, year = fetch_nib_historisk_ortofoto(expanded_bbox)
        if img:
            result["image"] = img
            result["source"] = src
            result["year"] = year
            result["log"].append(f"✅ Historisk ortofoto: {src} (år {year})")
            return result
        else:
            result["log"].append(f"⚠️ {src}")
    except Exception as e:
        result["log"].append(f"⚠️ Norge i Bilder feilet: {str(e)[:80]}")

    # Source 2: Geodata Online GeomapBilder variants (not truly historical, but different vintage)
    if _GDO_TOKEN_OK:
        try:
            from geodata_client import GEOMAP_BILDER_HISTORICAL_CANDIDATES
            for service, label in GEOMAP_BILDER_HISTORICAL_CANDIDATES:
                try:
                    img_bytes = _gdo._image_export(service, expanded_bbox, width=1200, height=1200)
                    if img_bytes and len(img_bytes) > 2000:
                        img = Image.open(io.BytesIO(img_bytes)).convert("RGB")
                        result["image"] = img
                        result["source"] = label
                        result["log"].append(f"✅ Historisk ortofoto (fallback): {label}")
                        return result
                except Exception:
                    continue
            result["log"].append("⚠️ Geodata Online historisk: kandidattjenester ga ingen data")
        except ImportError:
            pass

    result["log"].append("❌ Historisk flyfoto: Ingen kilde tilgjengelig — bruk manuell opplasting")
    return result


def fetch_miljodir_grunnforurensning(x: float, y: float, radius_m: float = 500.0) -> list:
    """Miljødirektoratet — kjente forurensinglokaliteter i nærheten."""
    try:
        from pyproj import Transformer
        tr = Transformer.from_crs(25833, 4326, always_xy=True)
        lon, lat = tr.transform(x, y)
    except ImportError:
        return []

    try:
        resp = requests.get(
            f"https://grfreg-api.miljodirektoratet.no/api/Lokalitet/hentNaereLokaliteter",
            params={"latitude": lat, "longitude": lon, "radius": int(radius_m)},
            timeout=8,
        )
        if resp.status_code == 200:
            data = resp.json()
            sites = data if isinstance(data, list) else data.get("lokaliteter", data.get("features", []))
            result = []
            for site in sites[:20]:
                props = site.get("properties", site) if isinstance(site, dict) else {}
                result.append({
                    "navn": props.get("navn") or props.get("lokalitetsnavn") or "Ukjent",
                    "status": props.get("status") or props.get("aktivitetsstatus") or "",
                    "type": props.get("type") or props.get("forurensningstype") or "",
                    "id": props.get("id") or props.get("lokalitetsId") or "",
                })
            return result
    except Exception:
        pass
    return []


def fetch_all_geodata(adresse: str, kommune: str, gnr: str, bnr: str,
                      radius_m: float = 300.0) -> dict:
    """Fetch all available geodata for a location. Returns dict with maps and data."""
    result = {
        "ortofoto": None, "ortofoto_source": "",
        "historisk": None, "historisk_source": "", "historisk_year": None,
        "losmasser": None, "berggrunn": None, "radon": None,
        "flom": None, "kvikkleire": None, "skred": None,
        "geologi_gdo": None,
        "forurensning_gdo": None, "kulturminner_gdo": None, "samfunnssikkerhet_gdo": None,
        "grunnforurensning": [],
        "geology_context": None,
        "coords": None, "bbox": None,
        "errors": [], "log": [],
    }

    # Step 1: Geocode — try multiple methods
    coords = _coords_from_address(adresse, kommune, gnr, bnr)

    # Primary method for Geodata Online users: fetch_tomt_polygon (same as Mulighetsstudie)
    site_polygon = None
    if _GDO_TOKEN_OK and gnr and bnr and kommune:
        try:
            knr = _resolve_kommunenummer(kommune)
            if knr:
                poly, msg = _gdo.fetch_tomt_polygon(knr, [(str(gnr).strip(), str(bnr).strip())])
                if poly:
                    site_polygon = poly
                    # Derive coordinates from polygon centroid
                    centroid = poly.centroid
                    coords = (centroid.x, centroid.y)
                    result["log"].append(f"✅ Tomt hentet via Geodata Online: {msg}")
                    result["log"].append(f"   Tomteareal: {int(poly.area)} m²")
                else:
                    result["log"].append(f"⚠️ Geodata Online tomt: {msg}")
        except Exception as e:
            result["log"].append(f"⚠️ Geodata Online tomt: {str(e)[:60]}")

    # Fallback: Geodata Online address search
    if not coords and _GDO_TOKEN_OK:
        try:
            hits = _gdo.address_search(adresse or "", kommune or "")
            if hits:
                result["log"].append(f"Geokoding via Geodata Online: {hits[0].get('label', '')}")
                coords = (hits[0]["x"], hits[0]["y"])
        except Exception:
            pass

    if not coords:
        result["errors"].append("Kunne ikke geokode adressen. Sjekk adresse/kommune i Project Setup.")
        return result

    x, y = coords
    result["coords"] = (x, y)
    result["log"].append(f"Koordinater: Ø={x:.0f}, N={y:.0f} (EPSG:25833)")
    bbox = site_polygon.bounds if site_polygon else _bbox_from_center(x, y, radius_m)
    result["bbox"] = bbox

    # Step 2: Ortofoto — Geodata Online HD first (like Mulighetsstudie), then Kartverket
    if _GDO_TOKEN_OK:
        try:
            img, src = _gdo.fetch_ortofoto(bbox=bbox, buffer_m=80.0, width=1200, height=1200)
            if img:
                result["ortofoto"] = img
                result["ortofoto_source"] = src
                result["log"].append(f"✅ Ortofoto: {src}")
        except Exception as e:
            result["log"].append(f"⚠️ Geodata Online ortofoto feilet: {str(e)[:60]}")

    if result["ortofoto"] is None:
        img = fetch_kartverket_ortofoto(bbox)
        if img:
            result["ortofoto"] = img
            result["ortofoto_source"] = "Kartverket (Norge i Bilder)"
            result["log"].append("✅ Ortofoto: Kartverket WMS")
        else:
            result["log"].append("❌ Ortofoto: Ingen kilde tilgjengelig")

    # Step 2b: Historisk ortofoto — Geodata Online → Norge i Bilder prosjekter → Kartverket historiske kart
    hist_result = fetch_historical_ortofoto(bbox)
    if hist_result.get("image"):
        result["historisk"] = hist_result["image"]
        result["historisk_source"] = hist_result.get("source", "")
        result["historisk_year"] = hist_result.get("year")
    result["log"].extend(hist_result.get("log", []))

    # Step 3: Geodata Online — geologi temakart (MapServer export)
    if _GDO_TOKEN_OK:
        try:
            from geodata_client import GEOMAP_DOKGEOLOGI_MS
            img = _geodata_online_map(GEOMAP_DOKGEOLOGI_MS, bbox)
            if img:
                result["geologi_gdo"] = img
                result["log"].append("✅ Geologi temakart: Geodata Online DOK Geologi")
        except Exception as e:
            result["log"].append(f"⚠️ Geodata Online geologi kart: {str(e)[:60]}")

    # Step 3b: Geodata Online — forurensning, kulturminner, samfunnssikkerhet temakart
    if _GDO_TOKEN_OK:
        try:
            from geodata_client import (GEOMAP_DOKFORURENSNING_MS,
                                         GEOMAP_DOKKULTUREMINNER_MS,
                                         GEOMAP_DOKSAMFUNNSSIKKERHET_MS)
            for key, service, label in [
                ("forurensning_gdo", GEOMAP_DOKFORURENSNING_MS, "DOK Forurensning"),
                ("kulturminner_gdo", GEOMAP_DOKKULTUREMINNER_MS, "DOK Kulturminner"),
                ("samfunnssikkerhet_gdo", GEOMAP_DOKSAMFUNNSSIKKERHET_MS, "DOK Samfunnssikkerhet"),
            ]:
                try:
                    img = _geodata_online_map(service, bbox)
                    if img:
                        result[key] = img
                        result["log"].append(f"✅ {label}: Geodata Online")
                except Exception as e:
                    result["log"].append(f"⚠️ {label}: {str(e)[:60]}")
        except ImportError:
            result["log"].append("⚠️ DOK-temakart: importfeil i geodata_client")

    # Step 4: NGU — løsmasser, berggrunn, radon (open WMS)
    for key, func, label in [
        ("losmasser", fetch_ngu_losmasser, "NGU Løsmasser"),
        ("berggrunn", fetch_ngu_berggrunn, "NGU Berggrunn"),
        ("radon", fetch_ngu_radon, "NGU Radon"),
    ]:
        img = func(bbox)
        if img:
            result[key] = img
            result["log"].append(f"✅ {label}")
        else:
            result["log"].append(f"⚠️ {label}: Ingen data (kan være utenfor kartlagt område)")

    # Step 5: NVE — flom, kvikkleire, skred (open WMS)
    for key, func, label in [
        ("flom", fetch_nve_flom, "NVE Flom"),
        ("kvikkleire", fetch_nve_kvikkleire, "NVE Kvikkleire"),
        ("skred", fetch_nve_skred, "NVE Skred"),
    ]:
        img = func(bbox)
        if img:
            result[key] = img
            result["log"].append(f"✅ {label}")
        else:
            result["log"].append(f"⚠️ {label}: Ingen data")

    # Step 6: Miljødirektoratet — grunnforurensning
    result["grunnforurensning"] = fetch_miljodir_grunnforurensning(x, y, radius_m=500.0)
    result["log"].append(f"✅ Miljødirektoratet: {len(result['grunnforurensning'])} lokaliteter")

    # Step 7: Geodata Online geology context (attribute data)
    if _GDO_TOKEN_OK:
        try:
            geo_poly = site_polygon
            if geo_poly is None:
                from shapely.geometry import box
                geo_poly = box(*bbox)
            result["geology_context"] = _gdo.fetch_geology_context(geo_poly, buffer_m=100.0)
            n_feat = len((result["geology_context"] or {}).get("features", []))
            result["log"].append(f"✅ Geodata Online geologi kontekst: {n_feat} objekter")
        except Exception as e:
            result["log"].append(f"⚠️ Geodata Online geologi kontekst: {str(e)[:60]}")

    return result


def geodata_summary_text(gd: dict) -> str:
    """Build a text summary of geodata findings for the AI prompt."""
    lines = []

    if gd.get("ortofoto"):
        lines.append(f"Ortofoto hentet fra {gd.get('ortofoto_source', 'ukjent kilde')}.")

    if gd.get("historisk"):
        hist_src = gd.get('historisk_source', 'ukjent kilde')
        hist_year = gd.get('historisk_year')
        if hist_year:
            lines.append(f"Historisk flyfoto fra {hist_year} er hentet og vedlagt ({hist_src}). Analyser bildet for å identifisere tidligere arealbruk og potensielle forurensningskilder.")
        else:
            lines.append(f"Historisk kart/flyfoto er hentet fra {hist_src} og vedlagt for analyse av historisk arealbruk.")
    else:
        lines.append("Historisk flyfoto er ikke tilgjengelig automatisk. Vurder historisk arealbruk basert på andre tilgjengelige kilder.")

    map_names = {
        "losmasser": "NGU Løsmassekart (kvartærgeologi)",
        "berggrunn": "NGU Bergrunnskart",
        "radon": "NGU Radonaktsomhetskart",
        "flom": "NVE Flomsonekart",
        "kvikkleire": "NVE Kvikkleirekart",
        "skred": "NVE Skredaktsomhetskart",
        "geologi_gdo": "Geodata Online DOK Geologi temakart",
        "forurensning_gdo": "DOK Forurensningskart (Geodata Online) — viser kartlagte forurensingssoner, deponier og registrerte forurensingslokaliteter",
        "kulturminner_gdo": "DOK Kulturminnekart (Geodata Online) — viser fredede og verneverdige kulturminner/kulturmiljøer",
        "samfunnssikkerhet_gdo": "DOK Samfunnssikkerhetskart (Geodata Online) — viser risiko- og sårbarhetsdata",
    }
    fetched = [label for key, label in map_names.items() if gd.get(key) is not None]
    missing = [label for key, label in map_names.items() if gd.get(key) is None]
    if fetched:
        lines.append(f"Følgende temakart er hentet og vedlagt som bilder for visuell analyse: {', '.join(fetched)}.")
    if missing:
        lines.append(f"Følgende temakart ga ingen data (trolig utenfor kartleggingsområde): {', '.join(missing)}.")

    contamination = gd.get("grunnforurensning", [])
    if contamination:
        lines.append(f"Miljødirektoratets database viser {len(contamination)} kjente forurensinglokaliteter innenfor 500 m:")
        for site in contamination[:8]:
            lines.append(f"  - {site.get('navn', 'Ukjent')}: {site.get('status', '')} ({site.get('type', '')})")
    else:
        lines.append("Miljødirektoratets grunnforurensningsdatabase: Ingen kjente forurensinglokaliteter funnet innenfor 500 m radius.")

    geo_ctx = gd.get("geology_context")
    if geo_ctx and geo_ctx.get("features"):
        lines.append(f"Geodata Online DOK Geologi: {len(geo_ctx['features'])} objekter funnet i nærheten.")
        for feat in geo_ctx["features"][:5]:
            attrs = feat.get("attributes", {})
            desc = " | ".join(f"{k}: {v}" for k, v in list(attrs.items())[:4] if v)
            if desc:
                lines.append(f"  - {desc}")

    return "\n".join(lines)

# --- 1. TEKNISK OPPSETT & GLOBALE STIER ---
st.set_page_config(page_title="Geo & Miljø (RIG-M) | Builtly", layout="wide", initial_sidebar_state="collapsed")

# Definerer stier som manglet (FIKSER NAMEERROR)
DB_DIR = Path("qa_database")
IMG_DIR = DB_DIR / "project_images"
SSOT_FILE = DB_DIR / "ssot.json"

# Oppretter mapper om de ikke finnes
DB_DIR.mkdir(exist_ok=True)
IMG_DIR.mkdir(exist_ok=True)

google_key = os.environ.get("GOOGLE_API_KEY")
if google_key:
    genai.configure(api_key=google_key)
else:
    st.error("Kritisk feil: Fant ingen API-nøkkel! Sjekk 'Environment Variables' i Render.")
    st.stop()


def render_html(html_string: str):
    st.markdown(html_string.replace('\n', ' '), unsafe_allow_html=True)

def logo_data_uri() -> str:
    for candidate in ["logo-white.png", "logo.png"]:
        if os.path.exists(candidate):
            suffix = Path(candidate).suffix.lower().replace(".", "") or "png"
            with open(candidate, "rb") as f:
                encoded = base64.b64encode(f.read()).decode("utf-8")
            return f"data:image/{suffix};base64,{encoded}"
    return ""

def find_page(base_name: str) -> str:
    for name in [base_name, base_name.lower(), base_name.capitalize()]:
        p = Path(f"pages/{name}.py")
        if p.exists(): return str(p)
    return ""

def clean_pdf_text(text):
    if text is None: return ""
    text = str(text)
    rep = {"–": "-", "—": "-", "“": '"', "”": '"', "‘": "'", "’": "'", "…": "...", "•": "-", "≤": "<=", "≥": ">="}
    for old, new in rep.items(): text = text.replace(old, new)
    return text.encode("latin-1", "replace").decode("latin-1")

def ironclad_text_formatter(text):
    text = clean_pdf_text(text)
    text = text.replace("$", "").replace("*", "").replace("_", "")
    text = re.sub(r"[-|=]{3,}", " ", text)
    text = re.sub(r"([^\s]{40})", r"\1 ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text

def nb_value(value):
    if value is None or value == "": return "-"
    if isinstance(value, float):
        if value.is_integer(): return str(int(value))
        if abs(value) >= 100: txt = f"{value:.0f}"
        elif abs(value) >= 10: txt = f"{value:.1f}"
        else: txt = f"{value:.3f}".rstrip("0").rstrip(".")
        return txt.replace(".", ",")
    if isinstance(value, int): return str(value)
    return clean_pdf_text(str(value))

def parse_numeric(value):
    if value is None: return None, None
    if isinstance(value, (int, float)): return float(value), None
    txt = str(value).strip()
    if not txt: return None, None
    qualifier = None
    txt = txt.replace(" ", "").replace(",", ".")
    low = txt.lower()
    if low in {"nd", "n.d.", "n.d", "na", "nan"}: return None, "nd"
    if txt.startswith("<"):
        qualifier = "<"
        txt = txt[1:]
    elif txt.startswith(">"):
        qualifier = ">"
        txt = txt[1:]
    txt = txt.replace("mg/kg", "")
    try: return float(txt), qualifier
    except Exception: return None, qualifier

def strip_empty_edges(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty: return df
    df2 = df.copy()
    df2 = df2.dropna(axis=0, how="all")
    df2 = df2.dropna(axis=1, how="all")
    return df2

ANALYTE_COLUMN_MAP = {
    "TOC (%)": 7, "As": 8, "Pb": 9, "Cd": 12, "Cr (tot)": 13, "Cu": 14, "Hg": 17, 
    "Ni": 18, "Zn": 19, "Bensen": 22, "Toluen": 23, "Etylbensen": 24, "Xylener": 25, 
    "C10-C12": 26, "C12-C35": 28, "Sum 16": 30, "B(a)p": 34, "Beskrivelse": 35,
}

DISPLAY_ANALYTES = ["As", "Pb", "Ni", "Zn", "C12-C35", "Sum 16", "B(a)p"]
CLASS_ORDER = {"TK1": 1, "TK2": 2, "TK3": 3, "TK4": 4, "TK5": 5, "TK>5": 6}
CLASS_FILL = {
    "TK1": (214, 236, 255), "TK2": (196, 235, 176), "TK3": (255, 242, 153),
    "TK4": (255, 202, 128), "TK5": (255, 153, 153), "TK>5": (232, 97, 97),
}

def get_font(size: int, bold: bool = False):
    candidates = ["/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"]
    for path in candidates:
        if os.path.exists(path):
            try: return ImageFont.truetype(path, size=size)
            except Exception: pass
    return ImageFont.load_default()

def wrap_text_px(text: str, font, max_width: int):
    text = clean_pdf_text(text)
    if not text: return [""]
    words = text.split()
    if not words: return [""]
    lines, current = [], words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if font.getbbox(candidate)[2] <= max_width: current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    final_lines = []
    for line in lines:
        while font.getbbox(line)[2] > max_width and len(line) > 10:
            cut = max(8, len(line) // 2)
            probe = line[:cut]
            while cut > 8 and font.getbbox(probe + "...")[2] > max_width:
                cut -= 1
                probe = line[:cut]
            final_lines.append(probe + "...")
            line = line[cut:]
        final_lines.append(line)
    return final_lines or [""]

def class_rank(class_code: str) -> int:
    return CLASS_ORDER.get(class_code or "", 0)

def classify_value(value, analyte: str, thresholds: dict):
    if analyte not in thresholds.get("TK1", {}): return None
    num, qualifier = parse_numeric(value)
    if num is None: return None
    tk1, tk2, tk3, tk4, tk5 = [thresholds[tk].get(analyte) if tk in thresholds else None for tk in ["TK1", "TK2", "TK3", "TK4", "TK5"]]
    if tk1 is None: return None
    if qualifier == "<" and num <= tk1: return "TK1"
    if num <= tk1: return "TK1"
    if tk2 is not None and num <= tk2: return "TK2"
    if tk3 is not None and num <= tk3: return "TK3"
    if tk4 is not None and num <= tk4: return "TK4"
    if tk5 is not None and num <= tk5: return "TK5"
    return "TK>5"

def split_dataframe(df: pd.DataFrame, chunk_size: int):
    if df is None or df.empty: return []
    return [df.iloc[start:start + chunk_size].reset_index(drop=True) for start in range(0, len(df), chunk_size)]

# --- 4. BILDER OG TABELL RENDERER ---
def render_table_image(df: pd.DataFrame, title: str, subtitle: str = "", row_class_column: str = None, cell_fill_lookup: dict = None, note: str = ""):
    df = df.copy().fillna("")
    title, subtitle, note = clean_pdf_text(title), clean_pdf_text(subtitle), clean_pdf_text(note)

    font_title, font_subtitle = get_font(34, bold=True), get_font(18, bold=False)
    font_header, font_body = get_font(18, bold=True), get_font(17, bold=False)

    side_pad, top_pad, cell_pad_x, cell_pad_y, table_width = 28, 24, 10, 9, 1520

    width_weights = []
    for col in df.columns:
        col_txt = str(col)
        if col_txt in {"Prøvepunkt", "Dybde", "Dybde (m)", "Fil", "Ark", "Høyeste klasse", "Styrende parameter"}: width_weights.append(1.0)
        elif col_txt in DISPLAY_ANALYTES or col_txt in {"Styrende verdi", "Klasse"}: width_weights.append(0.9)
        elif "Beskrivelse" in col_txt or "Kommentar" in col_txt: width_weights.append(2.8)
        elif col_txt in {"Datatype", "Innhold"}: width_weights.append(1.8)
        else: width_weights.append(1.3)
        
    total_weight = sum(width_weights) or 1
    col_widths = [max(95, int(table_width * w / total_weight)) for w in width_weights]

    header_height = 0
    header_wrapped = {}
    for col, width in zip(df.columns, col_widths):
        wrapped = wrap_text_px(str(col), font_header, width - (cell_pad_x * 2))
        header_wrapped[col] = wrapped
        header_height = max(header_height, len(wrapped) * 24 + (cell_pad_y * 2))

    row_heights, wrapped_cells = [], []
    for ridx in range(len(df)):
        row = df.iloc[ridx]
        row_wrap, row_height = {}, 0
        for col, width in zip(df.columns, col_widths):
            wrapped = wrap_text_px(str(row[col]), font_body, width - (cell_pad_x * 2))
            row_wrap[col] = wrapped
            row_height = max(row_height, len(wrapped) * 22 + (cell_pad_y * 2))
        row_heights.append(max(34, row_height))
        wrapped_cells.append(row_wrap)

    title_height = 66
    subtitle_height = 26 if subtitle else 0
    note_height = 32 if note else 0
    total_height = top_pad + title_height + subtitle_height + 14 + header_height + sum(row_heights) + note_height + 28
    
    image_width, image_height = table_width + side_pad * 2, total_height + 10
    img = Image.new("RGB", (image_width, image_height), (255, 255, 255))
    draw = ImageDraw.Draw(img)

    band_fill, header_fill, alt_fill = (236, 240, 245), (46, 62, 84), (248, 250, 252)
    grid_fill, title_fill, subtitle_fill, text_fill = (205, 212, 220), (29, 45, 68), (96, 108, 122), (35, 38, 43)

    draw.rounded_rectangle((12, 12, image_width - 12, image_height - 12), radius=18, outline=(219, 225, 232), width=2, fill=(255, 255, 255))
    draw.rounded_rectangle((18, 18, image_width - 18, 18 + title_height + subtitle_height + 10), radius=16, fill=band_fill)
    draw.text((side_pad, 28), title, font=font_title, fill=title_fill)
    if subtitle: draw.text((side_pad, 28 + 40), subtitle, font=font_subtitle, fill=subtitle_fill)

    x, y = side_pad, top_pad + title_height + subtitle_height + 10
    for col, width in zip(df.columns, col_widths):
        draw.rectangle((x, y, x + width, y + header_height), fill=header_fill)
        yy = y + cell_pad_y
        for line in header_wrapped[col]:
            draw.text((x + cell_pad_x, yy), clean_pdf_text(line), font=font_header, fill=(255, 255, 255))
            yy += 24
        x += width
    draw.rectangle((side_pad, y, side_pad + sum(col_widths), y + header_height), outline=grid_fill, width=1)

    y += header_height
    for ridx in range(len(df)):
        row = df.iloc[ridx]
        base_fill = alt_fill if ridx % 2 else (255, 255, 255)
        if row_class_column and row_class_column in row and str(row[row_class_column]) in CLASS_FILL:
            rf = CLASS_FILL[str(row[row_class_column])]
            base_fill = tuple(int((c + 255 * 3) / 4) for c in rf)
        x, row_height = side_pad, row_heights[ridx]
        for col, width in zip(df.columns, col_widths):
            cell_fill = cell_fill_lookup.get((ridx, str(col)), base_fill) if cell_fill_lookup else base_fill
            draw.rectangle((x, y, x + width, y + row_height), fill=cell_fill, outline=grid_fill, width=1)
            yy = y + cell_pad_y
            for line in wrapped_cells[ridx][col]:
                draw.text((x + cell_pad_x, yy), clean_pdf_text(line), font=font_body, fill=text_fill)
                yy += 22
            x += width
        y += row_height

    if note: draw.text((side_pad, y + 8), note, font=font_subtitle, fill=subtitle_fill)
    return img


def save_temp_image(img: Image.Image, suffix: str = ".png"):
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    img.save(tmp.name)
    tmp.close()
    return tmp.name

# --- 5. DATA EXTRACTION LOGIC ---
def read_generic_table(file_name: str, file_bytes: bytes):
    try:
        df = pd.read_csv(io.BytesIO(file_bytes)) if file_name.lower().endswith(".csv") else pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
        df = strip_empty_edges(df)
        if df.empty: return None
        df = df.head(30).copy()
        df.columns = [clean_pdf_text(str(c)) for c in df.columns]
        return df
    except Exception:
        return None

def is_multiconsult_summary_sheet(ws) -> bool:
    checks = [clean_pdf_text(ws.cell(18, 1).value), clean_pdf_text(ws.cell(19, 8).value), clean_pdf_text(ws.cell(20, 28).value)]
    haystack = " ".join(checks)
    return "Prøvepunkt" in haystack and "TUNGMETALLER" in haystack and "C12-C35" in haystack

def extract_metadata_lines(ws):
    lines = []
    for row in range(1, 11):
        vals = [clean_pdf_text(ws.cell(row, col).value) for col in [1, 21] if ws.cell(row, col).value]
        if vals: lines.append(" | ".join(vals))
    return lines

def extract_multiconsult_summary(file_name: str, file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    target_sheet = next((ws for ws in wb.worksheets if is_multiconsult_summary_sheet(ws)), None)
    if not target_sheet: return None

    ws, rows, thresholds, metadata_lines = target_sheet, [], {}, extract_metadata_lines(target_sheet)
    current_sample, table_started = None, False

    for r in range(1, min(ws.max_row, 220) + 1):
        a_txt = clean_pdf_text(ws.cell(r, 1).value).strip()
        e_val = ws.cell(r, 5).value

        if a_txt == "Prøvepunkt":
            table_started = True
            continue
        if not table_started: continue
        if a_txt == "Analyse" and rows: break
        
        if a_txt.startswith("Tilstandsklasse"):
            if match := re.search(r"(\d)", a_txt):
                tk = f"TK{match.group(1)}"
                thresholds[tk] = {}
                for analyte, col_idx in ANALYTE_COLUMN_MAP.items():
                    if analyte not in {"TOC (%)", "Beskrivelse"}:
                        val, _ = parse_numeric(ws.cell(r, col_idx).value)
                        thresholds[tk][analyte] = val
            continue
            
        if re.match(r"^SK\d+", a_txt): current_sample = a_txt
        if current_sample and e_val:
            row = {
                "Fil": clean_pdf_text(file_name), "Ark": clean_pdf_text(ws.title),
                "Prøvepunkt": current_sample, "Dybde (m)": clean_pdf_text(e_val),
                "Beskrivelse": clean_pdf_text(ws.cell(r, ANALYTE_COLUMN_MAP["Beskrivelse"]).value)
            }
            any_result = False
            for analyte, col_idx in ANALYTE_COLUMN_MAP.items():
                if analyte == "Beskrivelse": continue
                value = ws.cell(r, col_idx).value
                row[analyte] = value
                if value not in (None, ""): any_result = True
            if any_result: rows.append(row)

    if not rows: return None

    for row in rows:
        row.update({"class_rank": 0, "Høyeste klasse": "-", "Styrende parameter": "-", "Styrende verdi": "-", "_cell_classes": {}})
        for analyte in DISPLAY_ANALYTES:
            ccode = classify_value(row.get(analyte), analyte, thresholds)
            row["_cell_classes"][analyte] = ccode
            if class_rank(ccode) > row["class_rank"]:
                row.update({"class_rank": class_rank(ccode), "Høyeste klasse": ccode, "Styrende parameter": analyte, "Styrende verdi": nb_value(row.get(analyte))})
        if row["Høyeste klasse"] == "-" and row.get("Beskrivelse"):
            row.update({"Høyeste klasse": "TK1", "class_rank": 1})

    detail_df = pd.DataFrame(rows)
    detail_df["class_rank"] = detail_df["class_rank"].fillna(0).astype(int)
    
    summary_rows = []
    for sample, grp in detail_df.groupby("Prøvepunkt", sort=True):
        best = grp.sort_values(["class_rank"], ascending=[False]).iloc[0]
        summary_rows.append({k: best[k] for k in ["Prøvepunkt", "Høyeste klasse", "Dybde (m)", "Styrende parameter", "Styrende verdi", "Beskrivelse", "class_rank"]})
        
    summary_df = pd.DataFrame(summary_rows).sort_values(["class_rank", "Prøvepunkt"], ascending=[False, True]).reset_index(drop=True)
    exceedance_df = summary_df[summary_df["class_rank"] >= 2].copy()
    if exceedance_df.empty: exceedance_df = summary_df.head(8).copy()
    
    excerpt_df = detail_df[["Prøvepunkt", "Dybde (m)", "Beskrivelse"] + DISPLAY_ANALYTES + ["Høyeste klasse"]].copy()
    for col in DISPLAY_ANALYTES: excerpt_df[col] = excerpt_df[col].map(nb_value)

    cell_fill_lookup = {}
    for ridx, row in detail_df.reset_index(drop=True).iterrows():
        for analyte in DISPLAY_ANALYTES:
            if ccode := row["_cell_classes"].get(analyte):
                if ccode in CLASS_FILL: cell_fill_lookup[(ridx, analyte)] = CLASS_FILL[ccode]
        if (ccode := row.get("Høyeste klasse")) in CLASS_FILL:
            cell_fill_lookup[(ridx, "Høyeste klasse")] = CLASS_FILL[ccode]

    threshold_df = pd.DataFrame([{"Klasse": tk, **{k: nb_value(thresholds[tk].get(k)) for k in DISPLAY_ANALYTES}} for tk in ["TK1", "TK2", "TK3", "TK4", "TK5"] if tk in thresholds])
    counts = {tk: int((summary_df["Høyeste klasse"] == tk).sum()) for tk in ["TK1", "TK2", "TK3", "TK4", "TK5", "TK>5"]}

    prompt_lines = [f"KILDE: {file_name} | Ark: {ws.title}"] + metadata_lines[:8] + [
        f"Antall delprøver i analysetabellen: {len(detail_df)}", f"Antall prøvepunkt i analysetabellen: {summary_df['Prøvepunkt'].nunique()}", "Høyeste registrerte klasser per prøvepunkt:"
    ] + [f"- {r['Prøvepunkt']} | Dybde {r['Dybde (m)']} | {r['Styrende parameter']} = {r['Styrende verdi']} | {r['Høyeste klasse']} | {r['Beskrivelse']}" for _, r in summary_df.head(12).iterrows()]

    if not threshold_df.empty:
        prompt_lines.append("Tilstandsklassegrenser (utdrag):")
        for _, r in threshold_df.iterrows(): prompt_lines.append(f"- {r['Klasse']}: As {r['As']}, Pb {r['Pb']}, Ni {r['Ni']}, Zn {r['Zn']}, C12-C35 {r['C12-C35']}, Sum16 {r['Sum 16']}, B(a)p {r['B(a)p']}")

    source_overview_df = pd.DataFrame([{"Fil": clean_pdf_text(file_name), "Datatype": "Miljøteknisk analysetabell", "Ark": clean_pdf_text(ws.title), "Innhold": f"{summary_df['Prøvepunkt'].nunique()} prøvepunkt / {len(detail_df)} delprøver", "Kommentar": "Gjenkjent Multiconsult-oppsett med tilstandsklassegrenser og massebeskrivelser"}])

    return {"type": "multiconsult_summary", "prompt_text": "\n".join(prompt_lines), "source_overview_df": source_overview_df, "detail_df": detail_df, "sample_summary_df": summary_df.drop(columns=["class_rank"], errors="ignore"), "exceedance_df": exceedance_df.drop(columns=["class_rank"], errors="ignore"), "excerpt_df": excerpt_df, "threshold_df": threshold_df, "counts": counts, "cell_fill_lookup": cell_fill_lookup, "metadata_lines": metadata_lines}

def extract_drill_data(files):
    if not files: return {"prompt_text": "Ingen Excel/CSV-data ble lastet opp.", "source_overview_df": pd.DataFrame(), "sample_summary_df": pd.DataFrame(), "exceedance_df": pd.DataFrame(), "excerpt_df": pd.DataFrame(), "threshold_df": pd.DataFrame(), "counts": {}, "cell_fill_lookup": {}, "metadata_lines": []}
    
    prompt_parts, source_overview, sample_summaries, exceedances, excerpts, thresholds, counts, cell_fill_lookup, metadata_lines, excerpt_offset = [], [], [], [], [], [], {tk: 0 for tk in ["TK1", "TK2", "TK3", "TK4", "TK5", "TK>5"]}, {}, [], 0

    for f in files:
        file_name, file_bytes = clean_pdf_text(f.name), f.getvalue() if hasattr(f, "getvalue") else f.read()
        extracted = extract_multiconsult_summary(file_name, file_bytes) if file_name.lower().endswith((".xlsx", ".xlsm", ".xls")) else None

        if extracted:
            prompt_parts.append(extracted["prompt_text"])
            if not extracted["source_overview_df"].empty: source_overview.append(extracted["source_overview_df"])
            if not extracted["sample_summary_df"].empty: sample_summaries.append(extracted["sample_summary_df"])
            if not extracted["exceedance_df"].empty: exceedances.append(extracted["exceedance_df"])
            if not extracted["excerpt_df"].empty:
                excerpts.append(extracted["excerpt_df"])
                for (ridx, col), fill in extracted["cell_fill_lookup"].items(): cell_fill_lookup[(excerpt_offset + ridx, col)] = fill
                excerpt_offset += len(extracted["excerpt_df"])
            if not extracted["threshold_df"].empty: thresholds.append(extracted["threshold_df"])
            metadata_lines.extend(extracted.get("metadata_lines", []))
            for tk, val in extracted.get("counts", {}).items(): counts[tk] = counts.get(tk, 0) + int(val)
            continue

        generic_df = read_generic_table(file_name, file_bytes)
        if generic_df is not None:
            prompt_parts.append(f"KILDE: {file_name}\n{generic_df.head(20).to_csv(index=False, sep=';')}")
            source_overview.append(pd.DataFrame([{"Fil": file_name, "Datatype": "Generisk tabellfil", "Ark": "Første ark", "Innhold": f"{len(generic_df)} rader / {len(generic_df.columns)} kolonner", "Kommentar": "Ikke gjenkjent som standard analysematrise - vist som strukturert utdrag"}]))
            preview = generic_df.head(15).copy()
            if len(preview.columns) > 8: preview = preview.iloc[:, :8]
            preview.insert(0, "Fil", file_name)
            excerpts.append(preview)
            excerpt_offset += len(preview)
        else:
            prompt_parts.append(f"KILDE: {file_name}\n[Kunne ikke lese filinnholdet strukturert]")

    return {
        "prompt_text": "\n\n".join(part for part in prompt_parts if part) or "Ingen Excel/CSV-data ble lastet opp.",
        "source_overview_df": pd.concat(source_overview, ignore_index=True) if source_overview else pd.DataFrame(),
        "sample_summary_df": pd.concat(sample_summaries, ignore_index=True) if sample_summaries else pd.DataFrame(),
        "exceedance_df": pd.concat(exceedances, ignore_index=True) if exceedances else pd.DataFrame(),
        "excerpt_df": pd.concat(excerpts, ignore_index=True) if excerpts else pd.DataFrame(),
        "threshold_df": thresholds[0] if thresholds else pd.DataFrame(),
        "counts": counts, "cell_fill_lookup": cell_fill_lookup, "metadata_lines": metadata_lines
    }

# --- 6. DYNAMISK PDF MOTOR (CORPORATE LAYOUT) ---
def split_ai_sections(content: str):
    sections = []
    current = None
    for raw_line in content.splitlines():
        line = raw_line.strip()
        if line.startswith("#"):
            if current: sections.append(current)
            current = {"title": ironclad_text_formatter(line.lstrip("#").strip()), "lines": []}
            continue
        if current is None: 
            # SILER UT AI-INTRO: Ignorer alt før den første faktiske overskriften
            continue
        current["lines"].append(raw_line.rstrip())
    if current: sections.append(current)
    return sections

def is_subheading_line(line: str) -> bool:
    clean = line.strip()
    if not clean: return False
    if clean.startswith("##"): return True
    if clean.endswith(":") and len(clean) < 80 and len(clean.split()) <= 7: return True
    if clean == clean.upper() and any(ch.isalpha() for ch in clean) and len(clean) < 70: return True
    return False

def is_bullet_line(line: str) -> bool:
    return bool(re.match(r"^([-*•]|\d+\.)\s+", line.strip()))

def strip_bullet(line: str) -> str:
    return re.sub(r"^([-*•]|\d+\.)\s+", "", line.strip())

class BuiltlyCorporatePDF(FPDF):
    def header(self):
        if self.page_no() == 1: return
        self.set_y(11)
        self.set_text_color(88, 94, 102)
        self.set_font("Helvetica", "", 8)
        self.cell(0, 4, clean_pdf_text(self.header_left), 0, 0, "L")
        self.cell(0, 4, clean_pdf_text(self.header_right), 0, 1, "R")
        self.set_draw_color(188, 192, 197)
        self.line(18, 18, 192, 18)
        self.set_y(24)

    def footer(self):
        self.set_y(-12)
        self.set_draw_color(210, 214, 220)
        self.line(18, 285, 192, 285)
        self.set_font("Helvetica", "", 7)
        self.set_text_color(110, 114, 119)
        self.cell(60, 5, clean_pdf_text(self.doc_code), 0, 0, "L")
        self.cell(70, 5, clean_pdf_text("Utkast - krever faglig kontroll"), 0, 0, "C")
        self.cell(0, 5, clean_pdf_text(f"Side {self.page_no()}"), 0, 0, "R")

    def ensure_space(self, needed_height: float):
        if self.get_y() + needed_height > 272:
            self.add_page()

    def body_paragraph(self, text, first=False):
        text = ironclad_text_formatter(text)
        if not text: return
        self.set_x(20)
        self.set_font("Helvetica", "", 10.2 if not first else 10.6)
        self.set_text_color(35, 39, 43)
        self.multi_cell(170, 5.5 if not first else 5.7, text)
        self.ln(1.6)

    def subheading(self, text):
        text = ironclad_text_formatter(text.replace("##", "").rstrip(":"))
        self.ensure_space(28)
        self.ln(2)
        self.set_x(20)
        self.set_font("Helvetica", "B", 10.8)
        self.set_text_color(48, 64, 86)
        self.cell(0, 6, clean_pdf_text(text.upper()), 0, 1)
        self.set_draw_color(225, 229, 234)
        self.line(20, self.get_y(), 190, self.get_y())
        self.ln(2)

    def bullets(self, items, numbered=False):
        for idx, item in enumerate(items, start=1):
            clean = ironclad_text_formatter(item)
            if not clean: continue
            self.ensure_space(10)
            self.set_font("Helvetica", "", 10.1)
            self.set_text_color(35, 39, 43)
            start_y = self.get_y()
            self.set_xy(22, start_y)
            self.cell(6, 5.2, f"{idx}." if numbered else "-", 0, 0, "L")
            self.set_xy(28, start_y)
            self.multi_cell(162, 5.2, clean)
            self.ln(0.8)

    def section_title(self, title: str):
        self.ensure_space(50) # Sikrer plass til overskrift + minst noen linjer innhold (unngår ensomme overskrifter)
        self.ln(2)
        title = ironclad_text_formatter(title)
        num_match = re.match(r"^(\d+\.?\d*)\s*(.*)$", title)
        number, text = (num_match.group(1).rstrip("."), num_match.group(2).strip()) if num_match and (num_match.group(1).endswith(".") or num_match.group(2)) else (None, title)
        
        self.set_font("Helvetica", "B", 17)
        self.set_text_color(36, 50, 72)
        start_y = self.get_y()
        if number:
            self.set_xy(20, start_y)
            self.cell(12, 8, clean_pdf_text(number), 0, 0, "L")
            self.set_xy(34, start_y)
            self.multi_cell(156, 8, clean_pdf_text(text.upper()), 0, "L")
        else:
            self.set_xy(20, start_y)
            self.multi_cell(170, 8, clean_pdf_text(text.upper()), 0, "L")
        self.set_draw_color(204, 209, 216)
        self.line(20, self.get_y() + 1, 190, self.get_y() + 1)
        self.ln(5)

    def rounded_rect(self, x, y, w, h, r, style="", corners="1234"):
        try: super().rounded_rect(x, y, w, h, r, style, corners)
        except Exception: self.rect(x, y, w, h, style if style in {"F", "FD", "DF"} else "")

    def kv_card(self, items, x=None, width=80, title=None):
        if x is None: x = self.get_x()
        height = 10 + (len(items) * 6.3) + (7 if title else 0)
        self.ensure_space(height + 3)
        start_y = self.get_y()
        self.set_fill_color(245, 247, 249)
        self.set_draw_color(214, 219, 225)
        self.rounded_rect(x, start_y, width, height, 4, "1234", "DF")
        yy = start_y + 5
        if title:
            self.set_xy(x + 4, yy)
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(48, 64, 86)
            self.cell(width - 8, 5, clean_pdf_text(title.upper()), 0, 1)
            yy += 7
        for label, value in items:
            self.set_xy(x + 4, yy)
            self.set_font("Helvetica", "B", 8.6)
            self.set_text_color(72, 79, 87)
            self.cell(28, 5, clean_pdf_text(label), 0, 0)
            self.set_font("Helvetica", "", 8.6)
            self.set_text_color(35, 39, 43)
            self.multi_cell(width - 34, 5, clean_pdf_text(value))
            yy = self.get_y() + 1
        self.set_y(max(self.get_y(), start_y + height))

    def highlight_box(self, title: str, items, fill=(245, 247, 250), accent=(50, 77, 106)):
        self.set_font("Helvetica", "", 10)
        
        # DYNAMISK HØYDEBEREGNING: Forhindrer at tekst "sklir ut" under boksen
        total_text_h = 0
        for item in items:
            w = self.get_string_width(clean_pdf_text(item))
            lines = int((w / 145)) + 1
            total_text_h += (lines * 5.5) + 2

        box_h = 14 + total_text_h
        self.ensure_space(box_h + 5)
        x, y = 20, self.get_y()
        
        self.set_fill_color(*fill)
        self.set_draw_color(217, 223, 230)
        self.rounded_rect(x, y, 170, box_h, 4, "1234", "DF")
        self.set_fill_color(*accent)
        self.rect(x, y, 3, box_h, "F")
        self.set_xy(x + 6, y + 4)
        self.set_font("Helvetica", "B", 10.5)
        self.set_text_color(*accent)
        self.cell(0, 5, clean_pdf_text(title.upper()), 0, 1)
        self.set_text_color(35, 39, 43)
        self.set_font("Helvetica", "", 10)
        
        yy = y + 10
        for item in items:
            self.set_xy(x + 8, yy)
            self.cell(5, 5, "-", 0, 0)
            self.multi_cell(154, 5.2, clean_pdf_text(item))
            yy = self.get_y() + 2
            
        self.set_y(y + box_h + 3)

    def stats_row(self, stats):
        if not stats: return
        self.ensure_space(26)
        box_w, gap, x0, y = 40, 3.3, 20, self.get_y()
        for idx, (label, value, class_code) in enumerate(stats):
            x = x0 + idx * (box_w + gap)
            self.set_fill_color(*CLASS_FILL.get(class_code, (245, 247, 249)))
            self.set_draw_color(216, 220, 226)
            self.rounded_rect(x, y, box_w, 20, 3, "1234", "DF")
            self.set_xy(x, y + 3)
            self.set_font("Helvetica", "B", 15)
            self.set_text_color(33, 39, 45)
            self.cell(box_w, 7, clean_pdf_text(str(value)), 0, 1, "C")
            self.set_x(x)
            self.set_font("Helvetica", "", 7.8)
            self.set_text_color(75, 80, 87)
            self.multi_cell(box_w, 4, clean_pdf_text(label), 0, "C")
        self.set_y(y + 24)

    def figure_image(self, image_path, width=82, caption=""):
        img = Image.open(image_path)
        height = width * (img.height / img.width)
        self.ensure_space(height + 15)
        x, y = self.get_x(), self.get_y()
        self.set_draw_color(219, 223, 228)
        self.rect(x, y, width, height)
        self.image(image_path, x=x, y=y, w=width)
        self.set_y(y + height + 2)
        if caption:
            self.set_x(x)
            self.set_font("Helvetica", "I", 7.7)
            self.set_text_color(104, 109, 116)
            self.multi_cell(width, 4, clean_pdf_text(caption), 0, "C")
        self.set_y(y + height + 10)

    def table_image(self, img_path, width=170, caption=""):
        img = Image.open(img_path)
        height = width * (img.height / img.width)
        self.ensure_space(height + 15)
        x, y = 20, self.get_y()
        self.image(img_path, x=x, y=y, w=width)
        self.set_y(y + height + 2)
        if caption:
            self.set_x(20)
            self.set_font("Helvetica", "I", 7.7)
            self.set_text_color(104, 109, 116)
            self.multi_cell(width, 4, clean_pdf_text(caption), 0, "L")
        self.ln(6)


def build_cover_page(pdf, project_data, client, recent_img, hist_img, source_text):
    pdf.add_page()
    
    # 1. Logo Top Right (Over streken / ingen strek nødvendig her)
    if os.path.exists("logo.png"):
        try: pdf.image("logo.png", x=150, y=15, w=40)
        except: pass

    # 2. Overskrift og Tittel
    pdf.set_xy(20, 45)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(100, 105, 110)
    pdf.cell(80, 6, clean_pdf_text("RAPPORT"), 0, 1, "L")

    pdf.set_x(20)
    pdf.set_font("Helvetica", "B", 34) # Stor forside-tittel!
    pdf.set_text_color(20, 28, 38)
    # Begrenser bredden til 95, ALIGN L for å unngå strekking/mellomroms-bug
    pdf.multi_cell(95, 12, clean_pdf_text(project_data.get("p_name", "Geo & Miljø")), 0, 'L')
    
    pdf.ln(4)
    pdf.set_x(20)
    pdf.set_font("Helvetica", "B", 14) # FET, som bedt om
    pdf.set_text_color(64, 68, 74)
    # ALIGN L fikser de stygge mellomrommene!
    pdf.multi_cell(95, 6.5, clean_pdf_text("Miljøteknisk grunnundersøkelse, geoteknisk vurdering og overordnet tiltaksplan"), 0, 'L')

    # 3. Faktaboks Høyre side (Låst til X=118, Y=45)
    pdf.set_xy(118, 45)
    meta_items = [
        ("Oppdragsgiver", client or "-"),
        ("Emne", "Geo & Miljø (RIG-M)"),
        ("Dato / revisjon", datetime.now().strftime("%d.%m.%Y") + " / 01"),
        ("Dokumentkode", "Builtly-RIGM-001"),
    ]
    pdf.kv_card(meta_items, x=118, width=72)

    # 4. Forsidebilde (Kun ett bilde, max skalert)
    img_paths = []
    if recent_img:
        try: img_paths.append((save_temp_image(recent_img.convert("RGB"), ".jpg"), f"Nyere ortofoto ({source_text})"))
        except: pass
    elif hist_img: 
        try: img_paths.append((save_temp_image(hist_img.convert("RGB"), ".jpg"), "Historisk flyfoto"))
        except: pass

    if img_paths:
        img_path, caption = img_paths[0]
        with Image.open(img_path) as tmp_img:
            aspect = tmp_img.height / max(tmp_img.width, 1)
        
        w = 170
        h = w * aspect
        if h > 130:
            h = 130
            w = h / aspect
        
        x = 20 + (170 - w) / 2
        y = max(pdf.get_y() + 15, 115) # Sørger for at den starter under tekst/bokser
        
        pdf.set_xy(x, y)
        pdf.figure_image(img_path, width=w, caption=caption)
    else:
        pdf.set_fill_color(244, 246, 248)
        pdf.set_draw_color(220, 224, 228)
        pdf.rounded_rect(20, 115, 170, 80, 4, "1234", "DF")
        pdf.set_xy(24, 146)
        pdf.set_font("Helvetica", "I", 12)
        pdf.set_text_color(112, 117, 123)
        pdf.multi_cell(160, 6, clean_pdf_text("Kartgrunnlag legges inn automatisk eller via manuell opplasting i modulen."), 0, "C")

    # 5. Ansvarsfraskrivelse (Alltid låst i bunn)
    pdf.set_xy(20, 255)
    pdf.set_font("Helvetica", "", 8.8)
    pdf.set_text_color(104, 109, 116)
    pdf.multi_cell(170, 4.5, clean_pdf_text("Rapporten er generert av Builtly RIG-M AI på bakgrunn av prosjektdata, opplastet laboratoriemateriale og tilgjengelig kartgrunnlag. Dokumentet er et arbeidsutkast og skal underlegges faglig kontroll før bruk i prosjektering, byggesak eller myndighetsdialog."))

def build_toc_page(pdf, include_appendices=False):
    pdf.add_page()
    pdf.section_title("INNHOLDSFORTEGNELSE")
    items = ["1. Sammendrag og konklusjon", "2. Innledning og prosjektbeskrivelse", "3. Grunnforhold og geologi", "4. Naturfare og risiko", "5. Miljøteknisk historikk og forurensningsstatus", "6. Utførte grunnundersøkelser", "7. Resultater: grunnforhold og forurensning", "8. Geotekniske vurderinger", "9. Tiltaksplan og massehåndtering"]
    if include_appendices: items.extend(["Vedlegg A. Sammenstilling av analyseresultater", "Vedlegg B. Tilstandsklassegrenser (utdrag)"])
    pdf.set_font("Helvetica", "", 10.5)
    pdf.set_text_color(45, 49, 55)
    for item in items:
        pdf.ensure_space(9)
        y = pdf.get_y()
        pdf.set_x(22)
        pdf.cell(0, 6, clean_pdf_text(item), 0, 0, "L")
        pdf.set_draw_color(225, 229, 234)
        pdf.line(22, y + 6, 188, y + 6)
        pdf.ln(8)
    pdf.ln(6)
    pdf.highlight_box("Dokumentoppsett", ["Rapporten er bygget med tydelig seksjonshierarki, figurtekster og dedikerte tabellvedlegg for laboratoriedata.", "Opplastede lab-data sammenstilles i egne analyseresultattabeller fremfor å ligge skjult som råtekst i brødteksten."])

def render_maps(pdf, recent_img, hist_img, source_text):
    paths = []
    if recent_img: paths.append((save_temp_image(recent_img.convert("RGB"), ".jpg"), f"Figur 1. Nyere ortofoto. Kilde: {source_text}"))
    if hist_img: paths.append((save_temp_image(hist_img.convert("RGB"), ".jpg"), "Figur 2. Historisk flyfoto"))
    if not paths:
        pdf.highlight_box("Kartgrunnlag", ["Ingen kart- eller flyfoto ble lagt ved i denne genereringen."])
        return
    
    widths, x_positions, max_height = [82, 82], [20, 108], 0
    for idx, (img_path, caption) in enumerate(paths[:2]):
        img = Image.open(img_path)
        max_height = max(max_height, widths[idx] * (img.height / img.width))
    
    # VIKTIG FIKS: Vi flytter hele seksjonen med overskrift og kart til ny side hvis det ikke er plass til begge
    pdf.ensure_space(max_height + 40) 
    start_y = pdf.get_y()
    for idx, (img_path, caption) in enumerate(paths[:2]):
        pdf.set_xy(x_positions[idx], start_y)
        pdf.figure_image(img_path, width=widths[idx], caption=caption)
    pdf.set_y(start_y + max_height + 14)

def render_ai_section_body(pdf, lines):
    paragraph_buffer, bullet_buffer, first_para, empty_line_count = [], [], True, 0

    def flush_paragraph():
        nonlocal paragraph_buffer, first_para
        if paragraph_buffer:
            text = " ".join(line.strip() for line in paragraph_buffer if line.strip())
            if text:
                pdf.body_paragraph(text, first=first_para)
                first_para = False
        paragraph_buffer = []

    def flush_bullets():
        nonlocal bullet_buffer
        if bullet_buffer:
            pdf.bullets([strip_bullet(item) for item in bullet_buffer], numbered=all(re.match(r"^\d+\.\s+", item.strip()) for item in bullet_buffer))
        bullet_buffer = []

    for raw_line in lines:
        line = raw_line.strip()
        if not line:
            flush_paragraph()
            flush_bullets()
            empty_line_count += 1
            if empty_line_count == 1: pdf.ln(3) # Unngå evig lange blanke avsnitt
            continue
        empty_line_count = 0
        if is_subheading_line(line):
            flush_paragraph()
            flush_bullets()
            pdf.subheading(line)
            continue
        if is_bullet_line(line):
            flush_paragraph()
            bullet_buffer.append(line)
            continue
        flush_bullets()
        paragraph_buffer.append(line)

    flush_paragraph()
    flush_bullets()

def build_lab_summary_texts(lab_package):
    counts = lab_package.get("counts", {})
    summary_items = []
    if counts:
        summary_items.append(f"Prøvepunkter i TK2 eller høyere: {sum([counts.get(tk, 0) for tk in ['TK2', 'TK3', 'TK4', 'TK5', 'TK>5']])}")
        summary_items.append(f"Prøvepunkter med høyeste nivå i TK3 eller høyere: {sum([counts.get(tk, 0) for tk in ['TK3', 'TK4', 'TK5', 'TK>5']])}")
        summary_items.append(f"Prøvepunkter med høyeste nivå i TK5 eller over: {counts.get('TK5', 0) + counts.get('TK>5', 0)}")
    if not lab_package.get("exceedance_df", pd.DataFrame()).empty:
        for _, row in lab_package["exceedance_df"].head(3).iterrows():
            summary_items.append(f"{row['Prøvepunkt']} ({row['Dybde (m)']} m): {row['Styrende parameter']} = {row['Styrende verdi']} ({row['Høyeste klasse']}).")
    return summary_items[:6]

def create_full_report_pdf(name, client, content, recent_img, hist_img, source_text, lab_package, project_data):
    pdf = BuiltlyCorporatePDF("P", "mm", "A4")
    pdf.set_auto_page_break(True, margin=22) 
    pdf.set_margins(18, 18, 18)
    pdf.header_left, pdf.header_right, pdf.doc_code = clean_pdf_text(project_data.get("p_name", name)), clean_pdf_text("Builtly | RIG-M"), clean_pdf_text("Builtly-RIGM-001")

    build_cover_page(pdf, project_data, client, recent_img, hist_img, source_text)
    build_toc_page(pdf, include_appendices=not lab_package.get("excerpt_df", pd.DataFrame()).empty)

    sections = split_ai_sections(content) or [{"title": "1. SAMMENDRAG OG KONKLUSJON", "lines": [content]}]

    rendered_intro_boxes = False
    
    pdf.add_page()
    for idx, section in enumerate(sections):
        title = section.get("title", "")
        
        # Hvis det er et vedlegg, tvang-start på ny side, ellers la det flyte med sikret avstand
        if title.startswith("Vedlegg"):
            pdf.add_page()
        elif idx > 0:
            if pdf.get_y() > 30: pdf.ln(6) # Pustepause mellom kapittel

        # Spesialhåndtering for Kapittel 3 (Kart) for å unngå splitsing av overskrift og bilde
        if title.startswith("3."):
            pdf.ensure_space(120) # Krever mer plass til både tittel og kart (LÅSING)

        pdf.section_title(title)

        if title.startswith("1.") and not rendered_intro_boxes:
            # Dynamisk plassering av side-by-side kort (Sikrer at den kun tegnes EN gang selv om AI er teit)
            pdf.ensure_space(55)
            start_y = pdf.get_y()
            pdf.kv_card([("Prosjekt", project_data.get("p_name", name)), ("Lokasjon", f"{project_data.get('adresse', '')}, {project_data.get('kommune', '')}".strip(", ")), ("Gnr/Bnr", f"{project_data.get('gnr', '-')}/{project_data.get('bnr', '-')}") , ("Byggtype", project_data.get("b_type", "-")), ("BTA", f"{project_data.get('bta', 0)} m2")], x=20, width=82, title="Prosjektgrunnlag")
            end_left = pdf.get_y()
            
            pdf.set_xy(108, start_y)
            pdf.kv_card([("Kartgrunnlag", "Nyere + historisk" if recent_img and hist_img else "Delvis kartgrunnlag" if recent_img or hist_img else "Ikke vedlagt"), ("Lab-data", "Opplastet" if not lab_package.get("source_overview_df", pd.DataFrame()).empty else "Ikke opplastet"), ("Regelverk", project_data.get("land", "Norge"))], x=108, width=82, title="Datagrunnlag")
            end_right = pdf.get_y()
            
            pdf.set_y(max(end_left, end_right) + 6)
            
            if summary_items := build_lab_summary_texts(lab_package):
                pdf.highlight_box("Nøkkelfunn fra lab-data", summary_items)
                pdf.ln(4)
                
            rendered_intro_boxes = True

        if title.startswith("3."): render_maps(pdf, recent_img, hist_img, source_text)

        # Geodata temakart for kapittel 3 eller 4
        if title.startswith("3.") or title.startswith("4."):
            gd = st.session_state.geodata_result
            if gd and not getattr(pdf, '_geodata_maps_rendered', False):
                pdf._geodata_maps_rendered = True
                geo_map_pairs = []
                for key, label in [("losmasser", "Løsmassekart (NGU)"), ("berggrunn", "Bergrunnskart (NGU)"),
                                   ("radon", "Radon aktsomhet (NGU)"), ("geologi_gdo", "Geologi (Geodata Online)"),
                                   ("flom", "Flomsoner (NVE)"), ("kvikkleire", "Kvikkleire (NVE)"),
                                   ("skred", "Skredfare (NVE)")]:
                    img = gd.get(key)
                    if img is not None:
                        pil_img = img.convert("RGB") if isinstance(img, Image.Image) else img
                        geo_map_pairs.append((pil_img, label))

                if geo_map_pairs:
                    pdf.ensure_space(30)
                    pdf.set_font("Helvetica", "B", 10)
                    pdf.set_text_color(35, 39, 43)
                    pdf.cell(0, 7, clean_pdf_text("Temakart fra offentlige geodatakilder"), 0, 1)
                    pdf.ln(2)

                    for i in range(0, len(geo_map_pairs), 2):
                        pair = geo_map_pairs[i:i + 2]
                        fig_num = 3 + i
                        pdf.ensure_space(90)
                        start_y = pdf.get_y()
                        for j, (img, caption) in enumerate(pair):
                            x = 20 + j * 88
                            path = save_temp_image(img, ".jpg")
                            pdf.set_xy(x, start_y)
                            pdf.figure_image(path, width=82, caption=f"Figur {fig_num + j}. {caption}")
                        pdf.set_y(start_y + 82 * (img.height / max(img.width, 1)) + 14)

        if title.startswith("4.") and not lab_package.get("source_overview_df", pd.DataFrame()).empty:
            # Låser overskrift til tabellbilde
            pdf.ensure_space(60)
            source_table = render_table_image(lab_package["source_overview_df"], title="Opplastet analysegrunnlag", subtitle="Maskinelt lest og strukturert for rapportering", note="Tabellen viser hvilke kilder som faktisk ligger til grunn for vurderingene i denne genereringen.")
            pdf.table_image(save_temp_image(source_table), width=170, caption="Tabell 1. Oversikt over opplastet lab- og tabellgrunnlag.")

        render_ai_section_body(pdf, section.get("lines", []))

        if title.startswith("5.") and not lab_package.get("sample_summary_df", pd.DataFrame()).empty:
            # Låser statistikk og overskrift
            pdf.ensure_space(120)
            pdf.stats_row([("TK1 / rene", lab_package.get("counts", {}).get("TK1", 0), "TK1"), ("TK2", lab_package.get("counts", {}).get("TK2", 0), "TK2"), ("TK3", lab_package.get("counts", {}).get("TK3", 0), "TK3"), ("TK4-5", sum([lab_package.get("counts", {}).get(tk, 0) for tk in ["TK4", "TK5", "TK>5"]]), "TK5")])

            if not lab_package.get("exceedance_df", pd.DataFrame()).empty:
                top_table = render_table_image(lab_package["exceedance_df"].head(12), title="Høyeste påviste nivå per prøvepunkt", subtitle="Styrende parameter og klassifisering", row_class_column="Høyeste klasse", note="Radfarge følger høyeste registrerte tilstandsklasse per prøvepunkt.")
                pdf.table_image(save_temp_image(top_table), width=170, caption="Tabell 2. Sammendrag av styrende funn i opplastet laboratoriedata.")

            if not (excerpt_df := lab_package.get("excerpt_df", pd.DataFrame())).empty:
                pdf.ensure_space(80)
                preview = excerpt_df.head(12).copy()
                preview_img = render_table_image(preview, title="Analyseresultater og massebeskrivelser (utdrag)", subtitle="Fremstilt som rapporttabell i stedet for rå tekstutskrift", row_class_column="Høyeste klasse" if "Høyeste klasse" in preview.columns else None, cell_fill_lookup={(ridx, col): fill for (ridx, col), fill in lab_package.get("cell_fill_lookup", {}).items() if ridx < len(preview)}, note="Celler med farge markerer klassifiserte analyseresultater for de mest styrende parameterne.")
                pdf.table_image(save_temp_image(preview_img), width=170, caption="Tabell 3. Laboratoriedata presentert i vedleggsformat med klassifiserte nøkkelparametere.")

    if not (excerpt_df := lab_package.get("excerpt_df", pd.DataFrame())).empty:
        chunks = split_dataframe(excerpt_df, 12)
        for idx, chunk in enumerate(chunks, start=1):
            pdf.add_page()
            pdf.section_title(f"Vedlegg A. Sammenstilling av analyseresultater ({idx}/{len(chunks)})")
            raw_img = render_table_image(chunk.reset_index(drop=True), title="Vedleggstabell - analyseresultater", subtitle="Opplastet lab-data i rapportvennlig vedleggsformat", row_class_column="Høyeste klasse" if "Høyeste klasse" in chunk.columns else None, cell_fill_lookup={(ridx - ((idx - 1) * 12), col): fill for (ridx, col), fill in {(r, c): f for (r, c), f in lab_package.get("cell_fill_lookup", {}).items() if r >= (idx - 1) * 12 and r < idx * 12}.items()}, note="Utvalgte analyttkolonner er beholdt for å gjøre vedlegget lesbart i A4-format.")
            pdf.table_image(save_temp_image(raw_img), width=170, caption=f"Vedlegg A{idx}. Strukturert tabellutdrag fra opplastet laboratoriedata.")

    if not (threshold_df := lab_package.get("threshold_df", pd.DataFrame())).empty:
        pdf.add_page()
        pdf.section_title("Vedlegg B. Tilstandsklassegrenser (utdrag)")
        threshold_img = render_table_image(threshold_df, title="Tilstandsklassegrenser brukt i klassifisering", subtitle="Utvalgte parametere fra opplastet vedlegg", row_class_column="Klasse", note="Grenseverdiene er brukt til å markere relevante analyttceller i tabellene over.")
        pdf.table_image(save_temp_image(threshold_img), width=165, caption="Vedlegg B. Utdrag av tilstandsklassegrenser for sentrale analyttgrupper.")

    out = pdf.output(dest="S")
    return bytes(out) if isinstance(out, (bytes, bytearray)) else out.encode("latin-1")

# --- 7. UI OG RESTERENDE KODE ---
st.markdown("<style>/* Skjuler Streamlit-branding */\n#MainMenu {visibility: hidden;}\nfooter {visibility: hidden;}\nheader {visibility: hidden;}</style>", unsafe_allow_html=True)
st.markdown(
    """
<style>
    :root {
        --bg: #06111a; --panel: rgba(10, 22, 35, 0.78);
        --stroke: rgba(120, 145, 170, 0.18); --text: #f5f7fb; --muted: #9fb0c3; --soft: #c8d3df;
        --accent: #38bdf8; --radius-lg: 16px; --radius-xl: 24px;
    }
    html, body, [class*="css"] { font-family: Inter, ui-sans-serif, system-ui, -apple-system, sans-serif; }
    .stApp { background-color: var(--bg) !important; color: var(--text); }
    header[data-testid="stHeader"] { visibility: hidden; height: 0; }
    .block-container { max-width: 1280px !important; padding-top: 1.5rem !important; padding-bottom: 4rem !important; }

    .brand-logo { height: 65px; filter: drop-shadow(0 0 18px rgba(120,220,225,0.08)); }

    .top-shell { margin-bottom: 2rem; display: flex; justify-content: space-between; align-items: center; }

    button[kind="primary"] { background: linear-gradient(135deg, rgba(56,194,201,0.96), rgba(120,220,225,0.96)) !important; color: #041018 !important; border: none !important; font-weight: 750 !important; border-radius: 12px !important; padding: 12px 24px !important; font-size: 1.05rem !important; transition: all 0.2s ease !important; }
    button[kind="primary"]:hover { transform: translateY(-2px) !important; box-shadow: 0 12px 24px rgba(56,194,201,0.25) !important; }
    button[kind="secondary"] { background-color: rgba(255,255,255,0.05) !important; color: #f8fafc !important; border: 1px solid rgba(120,145,170,0.3) !important; border-radius: 12px !important; font-weight: 650 !important; padding: 10px 24px !important; transition: all 0.2s; }
    button[kind="secondary"]:hover { background-color: rgba(56,194,201,0.1) !important; border-color: var(--accent) !important; color: var(--accent) !important; transform: translateY(-2px) !important;}

    div[data-baseweb="base-input"], div[data-baseweb="select"] > div, .stTextArea > div > div > div { background-color: #0d1824 !important; border: 1px solid rgba(120, 145, 170, 0.4) !important; border-radius: 8px !important; }
    .stTextInput input, .stNumberInput input, .stTextArea textarea, div[data-baseweb="select"] * { background-color: transparent !important; color: #ffffff !important; -webkit-text-fill-color: #ffffff !important; border: none !important; box-shadow: none !important; }
    .stTextInput input:focus, .stNumberInput input:focus, .stTextArea textarea:focus { border: none !important; }
    div[data-baseweb="base-input"]:focus-within, div[data-baseweb="select"] > div:focus-within, .stTextArea > div > div > div:focus-within { border-color: var(--accent) !important; box-shadow: 0 0 0 1px rgba(56, 194, 201, 0.5) !important; }
    ul[data-baseweb="menu"] { background-color: #0d1824 !important; border: 1px solid rgba(120, 145, 170, 0.4) !important; }
    ul[data-baseweb="menu"] li { color: #ffffff !important; -webkit-text-fill-color: #ffffff !important; }
    ul[data-baseweb="menu"] li:hover { background-color: rgba(56, 194, 201, 0.1) !important; }
    div[data-testid="InputInstructions"], div[data-testid="InputInstructions"] > span { color: #9fb0c3 !important; -webkit-text-fill-color: #9fb0c3 !important; }
    .stTextInput label, .stSelectbox label, .stNumberInput label, .stTextArea label, .stFileUploader label { color: #c8d3df !important; font-weight: 600 !important; font-size: 0.95rem !important; margin-bottom: 4px !important; }

    div[data-testid="stExpander"] details, div[data-testid="stExpander"] details summary, div[data-testid="stExpander"] { background-color: #0c1520 !important; color: #f5f7fb !important; border-radius: 12px !important; }
    div[data-testid="stExpander"] details summary:hover { background-color: rgba(255,255,255,0.03) !important; }
    div[data-testid="stExpander"] details summary p { color: #f5f7fb !important; font-weight: 650 !important; }
    div[data-testid="stExpander"] { border: 1px solid rgba(120,145,170,0.2) !important; margin-bottom: 1rem !important; }
    div[data-testid="stExpanderDetails"] { background: transparent !important; color: #f5f7fb !important; }
    div[data-testid="stExpanderDetails"] > div > div > div { background-color: transparent !important; }

    [data-testid="stFileUploaderDropzone"] { background-color: #0d1824 !important; border: 1px dashed rgba(120, 145, 170, 0.6) !important; border-radius: 12px !important; padding: 2rem !important; }
    [data-testid="stFileUploaderDropzone"]:hover { border-color: #38c2c9 !important; background-color: rgba(56, 194, 201, 0.05) !important; }
    [data-testid="stFileUploaderDropzone"] * { color: #c8d3df !important; }
    [data-testid="stFileUploaderFileData"] { background-color: rgba(255,255,255,0.02) !important; color: #f5f7fb !important; border-radius: 8px !important;}
    [data-testid="stAlert"] { background-color: rgba(56, 194, 201, 0.05) !important; border: 1px solid rgba(56, 194, 201, 0.2) !important; border-radius: 12px !important; }
    [data-testid="stAlert"] * { color: #f5f7fb !important; }
</style>
""",
    unsafe_allow_html=True,
)

# --- 8. SESSION STATE (UI) ---
if "project_data" not in st.session_state:
    st.session_state.project_data = {
        "p_name": "", "c_name": "", "p_desc": "", "adresse": "", "kommune": "", "gnr": "", "bnr": "", "b_type": "Næring", "etasjer": 1, "bta": 0, "land": "Norge"
    }

if "geo_maps" not in st.session_state:
    st.session_state.geo_maps = {"recent": None, "historical": None, "source": "Ikke hentet"}
if "geodata_result" not in st.session_state:
    st.session_state.geodata_result = None

if st.session_state.project_data.get("p_name") == "":
    if SSOT_FILE.exists():
        with open(SSOT_FILE, "r", encoding="utf-8") as f:
            st.session_state.project_data = json.load(f)

if st.session_state.project_data.get("p_name") in ["", "Nytt Prosjekt"]:
    logo_html = f'<img src="{logo_data_uri()}" class="brand-logo">' if logo_data_uri() else '<h2 style="margin:0; color:white;">Builtly</h2>'
    render_html(f"<div style='margin-bottom:2rem;'>{logo_html}</div>")
    st.warning("⚠️ **Handling kreves:** Du må sette opp prosjektdataen før du kan bruke denne modulen.")
    if find_page("Project"):
        if st.button("⚙️ Gå til Project Setup", type="primary"):
            st.switch_page(find_page("Project"))
    st.stop()

pd_state = st.session_state.project_data

# --- 9. HEADER ---
top_l, top_r = st.columns([4, 1])
with top_l:
    logo_html = f'<img src="{logo_data_uri()}" class="brand-logo">' if logo_data_uri() else '<h2 style="margin:0; color:white;">Builtly</h2>'
    render_html(logo_html)
with top_r:
    st.markdown("<div style='margin-top: 0.5rem;'></div>", unsafe_allow_html=True)
    if st.button("← Tilbake til SSOT", use_container_width=True, type="secondary"):
        st.switch_page(find_page("Project"))

st.markdown("<hr style='border-color: rgba(120,145,170,0.1); margin-top: -1rem; margin-bottom: 2rem;'>", unsafe_allow_html=True)

# --- 10. KARTVERKET + GOOGLE MAPS FALLBACK ---
def fetch_kartverket_og_google(adresse, kommune, gnr, bnr, api_key):
    nord, ost = None, None
    adr_clean = adresse.replace(",", "").strip() if adresse else ""
    kom_clean = kommune.replace(",", "").strip() if kommune else ""

    queries = []
    if adr_clean and kom_clean:
        queries.append(f"{adr_clean} {kom_clean}")
    if adr_clean:
        queries.append(adr_clean)
    if gnr and bnr and kom_clean:
        queries.append(f"{kom_clean} {gnr}/{bnr}")

    for q in queries:
        safe_query = urllib.parse.quote(q)
        url = f"https://ws.geonorge.no/adresser/v1/sok?sok={safe_query}&fuzzy=true&utkoordsys=25833&treffPerSide=1"
        try:
            resp = requests.get(url, timeout=4)
            if resp.status_code == 200 and resp.json().get("adresser"):
                hit = resp.json()["adresser"][0]
                nord = hit.get("representasjonspunkt", {}).get("nord")
                ost = hit.get("representasjonspunkt", {}).get("øst")
                break
        except Exception:
            pass

    if nord and ost:
        min_x, max_x = float(ost) - 150, float(ost) + 150
        min_y, max_y = float(nord) - 150, float(nord) + 150
        url_orto = (
            "https://wms.geonorge.no/skwms1/wms.nib?service=WMS&request=GetMap&version=1.1.1"
            f"&layers=ortofoto&styles=&srs=EPSG:25833&bbox={min_x},{min_y},{max_x},{max_y}&width=800&height=800&format=image/png"
        )
        try:
            r1 = requests.get(url_orto, timeout=5)
            if r1.status_code == 200 and len(r1.content) > 5000:
                return Image.open(io.BytesIO(r1.content)).convert("RGB"), "Kartverket (Norge i Bilder)"
        except Exception:
            pass

    if api_key and (adr_clean or kom_clean):
        query = f"{adr_clean}, {kom_clean}, Norway"
        safe_query = urllib.parse.quote(query)
        url_gmaps = f"https://maps.googleapis.com/maps/api/staticmap?center={safe_query}&zoom=19&size=600x600&maptype=satellite&key={api_key}"
        try:
            r2 = requests.get(url_gmaps, timeout=5)
            if r2.status_code == 200:
                return Image.open(io.BytesIO(r2.content)).convert("RGB"), "Google Maps Satellite"
        except Exception:
            pass

    return None, "Kunne ikke hente kart."


# --- 11. UI FOR GEO MODUL ---
st.markdown("<h1 style='font-size: 2.5rem; margin-bottom: 0;'>🌍 Geo & Miljø (RIG-M)</h1>", unsafe_allow_html=True)
st.markdown("<p style='color: #9fb0c3; font-size: 1.1rem; margin-bottom: 2rem;'>AI-agent for miljøteknisk grunnundersøkelse og tiltaksplan.</p>", unsafe_allow_html=True)

st.success(f"✅ Prosjektdata for **{pd_state['p_name']}** er synkronisert fra Project SSOT.")

with st.expander("1. Prosjekt & Lokasjon (SSOT)", expanded=False):
    c1, c2 = st.columns(2)
    st.text_input("Prosjektnavn", value=pd_state["p_name"], disabled=True)
    st.text_input("Gnr/Bnr", value=f"{pd_state['gnr']} / {pd_state['bnr']}", disabled=True)

with st.expander("2. Geodata & Kartgrunnlag (Auto-henting)", expanded=True):
    st.markdown("Henter automatisk ortofoto, geologiske kart, naturfare og forurensningsdata fra åpne norske geodatakilder (NGU, NVE, Kartverket, Miljødirektoratet).")

    col_btn, col_status = st.columns([1, 2])
    with col_btn:
        if st.button("🌐 Hent alle geodata automatisk", type="primary", use_container_width=True):
            with st.spinner("Henter geodata fra NGU, NVE, Kartverket, Geodata Online, Miljødirektoratet..."):
                gd = fetch_all_geodata(pd_state["adresse"], pd_state["kommune"], pd_state["gnr"], pd_state["bnr"])
                st.session_state.geodata_result = gd
                if gd.get("ortofoto"):
                    st.session_state.geo_maps["recent"] = gd["ortofoto"]
                    st.session_state.geo_maps["source"] = gd.get("ortofoto_source", "Geodata")
                if gd.get("historisk"):
                    st.session_state.geo_maps["historical"] = gd["historisk"]
                if gd.get("errors"):
                    for err in gd["errors"]:
                        st.warning(f"⚠️ {err}")
                st.rerun()

    gd = st.session_state.geodata_result
    if gd:
        with col_status:
            n_maps = sum(1 for k in ["ortofoto", "historisk", "losmasser", "berggrunn", "radon", "flom", "kvikkleire", "skred", "geologi_gdo", "forurensning_gdo", "kulturminner_gdo", "samfunnssikkerhet_gdo"] if gd.get(k) is not None)
            n_contam = len(gd.get("grunnforurensning", []))
            hist_label = ""
            if gd.get("historisk"):
                hist_year = gd.get("historisk_year", "")
                hist_label = f" | Historisk flyfoto: {hist_year}" if hist_year else " | Historisk kart hentet"
            st.success(f"✅ {n_maps} kartlag hentet | {n_contam} forurensinglokaliteter{hist_label}")

        # --- Diagnostikk-logg ---
        if gd.get("log"):
            with st.popover("📋 Hentingslogg"):
                for line in gd["log"]:
                    st.text(line)

        if gd.get("errors"):
            for err in gd["errors"]:
                st.warning(f"⚠️ {err}")

        # --- Ortofoto (nyeste + historisk) ---
        has_recent = gd.get("ortofoto") is not None
        has_hist = gd.get("historisk") is not None

        if has_recent and has_hist:
            st.markdown("##### 📸 Ortofoto — Nyeste vs. Historisk")
            col_recent, col_hist = st.columns(2)
            with col_recent:
                st.image(gd["ortofoto"], caption=f"Nyeste ({gd.get('ortofoto_source', '')})", use_container_width=True)
            with col_hist:
                hist_year = gd.get("historisk_year", "")
                hist_caption = f"Historisk ({hist_year})" if hist_year else "Historisk"
                hist_caption += f" — {gd.get('historisk_source', '')}"
                st.image(gd["historisk"], caption=hist_caption, use_container_width=True)
        elif has_recent:
            st.markdown(f"##### 📸 Ortofoto — {gd.get('ortofoto_source', '')}")
            st.image(gd["ortofoto"], use_container_width=True)
        elif has_hist:
            hist_year = gd.get("historisk_year", "")
            st.markdown(f"##### 📸 Historisk flyfoto ({hist_year})")
            st.image(gd["historisk"], caption=gd.get("historisk_source", ""), use_container_width=True)

        # --- Geodata Online geologi ---
        if gd.get("geologi_gdo"):
            st.markdown("##### 🗺️ Geologi temakart (Geodata Online)")
            st.image(gd["geologi_gdo"], caption="DOK Geologi — Geodata Online", use_container_width=True)

        # --- NGU Temakart ---
        ngu_maps = []
        if gd.get("losmasser"):
            ngu_maps.append(("Løsmassekart (NGU)", gd["losmasser"]))
        if gd.get("berggrunn"):
            ngu_maps.append(("Bergrunnskart (NGU)", gd["berggrunn"]))
        if gd.get("radon"):
            ngu_maps.append(("Radon aktsomhet (NGU)", gd["radon"]))

        if ngu_maps:
            st.markdown("##### 🪨 NGU — Geologi & Radon")
            cols = st.columns(len(ngu_maps))
            for idx, (caption, img) in enumerate(ngu_maps):
                with cols[idx]:
                    st.image(img, caption=caption, use_container_width=True)

        # --- NVE Temakart ---
        nve_maps = []
        if gd.get("flom"):
            nve_maps.append(("Flomsoner (NVE)", gd["flom"]))
        if gd.get("kvikkleire"):
            nve_maps.append(("Kvikkleire (NVE)", gd["kvikkleire"]))
        if gd.get("skred"):
            nve_maps.append(("Skredfare (NVE)", gd["skred"]))

        if nve_maps:
            st.markdown("##### 🌊 NVE — Naturfare")
            cols = st.columns(len(nve_maps))
            for idx, (caption, img) in enumerate(nve_maps):
                with cols[idx]:
                    st.image(img, caption=caption, use_container_width=True)

        # --- Geodata Online DOK Temakart (forurensning, kulturminner, samfunnssikkerhet) ---
        dok_maps = []
        if gd.get("forurensning_gdo"):
            dok_maps.append(("Forurensning (DOK)", gd["forurensning_gdo"]))
        if gd.get("kulturminner_gdo"):
            dok_maps.append(("Kulturminner (DOK)", gd["kulturminner_gdo"]))
        if gd.get("samfunnssikkerhet_gdo"):
            dok_maps.append(("Samfunnssikkerhet (DOK)", gd["samfunnssikkerhet_gdo"]))

        if dok_maps:
            st.markdown("##### 🏛️ DOK — Forurensning, Kulturminner & Samfunnssikkerhet (Geodata Online)")
            cols = st.columns(len(dok_maps))
            for idx, (caption, img) in enumerate(dok_maps):
                with cols[idx]:
                    st.image(img, caption=caption, use_container_width=True)

        # --- Miljødirektoratet ---
        contam = gd.get("grunnforurensning", [])
        if contam:
            st.markdown("##### ☣️ Miljødirektoratet — Kjente forurensinglokaliteter (500 m radius)")
            for site in contam:
                status_icon = "🔴" if "aktiv" in (site.get("status", "")).lower() else "🟡"
                st.markdown(f"{status_icon} **{site.get('navn', 'Ukjent')}** — {site.get('status', '')} ({site.get('type', '')})")
        elif gd.get("coords"):
            st.info("✅ Ingen kjente forurensinglokaliteter funnet i Miljødirektoratets database (500 m radius).")

    # --- Fallback: manuell opplasting ---
    with st.popover("📎 Manuell opplasting (valgfritt)"):
        man_recent = st.file_uploader("Last opp nyere ortofoto", type=["png", "jpg", "jpeg"], key="man_ortofoto")
        if man_recent:
            st.session_state.geo_maps["recent"] = Image.open(man_recent).convert("RGB")
            st.session_state.geo_maps["source"] = "Manuelt opplastet"

        if st.session_state.geo_maps.get("historical"):
            st.info("✅ Historisk flyfoto er allerede hentet automatisk. Du kan overstyre med manuell opplasting nedenfor.")
        man_hist = st.file_uploader("Last opp historisk flyfoto (f.eks. 1950-tallet)", type=["png", "jpg", "jpeg"], key="man_hist")
        if man_hist:
            st.session_state.geo_maps["historical"] = Image.open(man_hist).convert("RGB")

with st.expander("3. Laboratoriedata & Plantegninger", expanded=True):
    st.info("Slipp Excel/CSV-filer med prøvesvar her. AI-en leser verdiene og tilstandsklassifiserer massene.")

    if "project_images" in st.session_state and len(st.session_state.project_images) > 0:
        st.success(f"📎 Auto-hentet {len(st.session_state.project_images)} arkitekttegninger fra Project Setup for vurdering av gravegrenser!")

    files = st.file_uploader("Last opp Excel/CSV med boreresultater:", accept_multiple_files=True, type=["xlsx", "csv", "xls"])

st.markdown("<br>", unsafe_allow_html=True)
if st.button("🚀 GENERER GEOTEKNISK & MILJØTEKNISK RAPPORT", type="primary", use_container_width=True):
    gd = st.session_state.geodata_result
    if not st.session_state.geo_maps["recent"] and not st.session_state.geo_maps["historical"] and not gd:
        st.error("🛑 **Stopp:** Du må hente geodata (Steg 2) eller laste opp kart manuelt før du kan generere rapport.")
        st.stop()

    with st.spinner("📊 Tolker lab-data, geodata, temakart og arkitekttegninger..."):
        lab_package = extract_drill_data(files) if files else extract_drill_data([])
        extracted_data = lab_package["prompt_text"] if files else "Ingen opplastet lab-data. Vurderingen baseres på visuell befaring, geodata og historikk."

        images_for_geo = []
        if st.session_state.geo_maps["recent"]:
            images_for_geo.append(st.session_state.geo_maps["recent"])
        if st.session_state.geo_maps["historical"]:
            images_for_geo.append(st.session_state.geo_maps["historical"])

        # Add all geodata thematic maps as images for AI analysis
        geodata_text = ""
        if gd:
            for key in ["losmasser", "berggrunn", "radon", "flom", "kvikkleire", "skred", "geologi_gdo", "forurensning_gdo", "kulturminner_gdo", "samfunnssikkerhet_gdo"]:
                img = gd.get(key)
                if img is not None:
                    images_for_geo.append(img if isinstance(img, Image.Image) else img.convert("RGB"))
            geodata_text = geodata_summary_text(gd)

        if "project_images" in st.session_state and isinstance(st.session_state.project_images, list):
            images_for_geo.extend(st.session_state.project_images)

        try:
            valid_models = [m.name for m in genai.list_models() if "generateContent" in m.supported_generation_methods]
            valgt_modell = valid_models[0]
            for fav in ["models/gemini-1.5-pro", "models/gemini-1.5-flash"]:
                if fav in valid_models:
                    valgt_modell = fav
                    break
        except Exception:
            st.error("Kunne ikke koble til Google AI.")
            st.stop()

        model = genai.GenerativeModel(valgt_modell)
        # Build historical imagery prompt text
        if st.session_state.geo_maps["historical"]:
            hist_year = (gd or {}).get("historisk_year", "")
            hist_source = (gd or {}).get("historisk_source", "")
            if hist_year:
                hist_tekst = f"Et historisk flyfoto fra {hist_year} er lagt ved ({hist_source}). Analyser dette for å identifisere tidligere arealbruk, industrivirksomhet, tankanlegg, deponier eller annen aktivitet som kan ha medført forurensning."
            else:
                hist_tekst = f"Et historisk kart/flyfoto er lagt ved ({hist_source}). Analyser dette for å identifisere historisk arealbruk."
        else:
            hist_tekst = "Historisk flyfoto mangler, gjør en kvalifisert antakelse basert på tilgjengelige data."

        prompt = f"""
        Du er Builtly RIG-M AI, en presis senior miljørådgiver og geotekniker.
        Skriv en formell, stram og troverdig "Miljøteknisk grunnundersøkelse og tiltaksplan for forurenset grunn" for:

        PROSJEKT: {pd_state['p_name']} ({pd_state['b_type']}, {pd_state['bta']} m2)
        LOKASJON: {pd_state['adresse']}, {pd_state['kommune']}. Gnr {pd_state['gnr']}/Bnr {pd_state['bnr']}.
        REGELVERK: {pd_state['land']}

        KUNDENS PROSJEKTNARRATIV: "{pd_state['p_desc']}"
        KARTSTATUS: {hist_tekst}

        GEODATA FRA OFFENTLIGE KILDER:
        {geodata_text if geodata_text else "Ingen geodata hentet. Vurder basert på generell kunnskap om området."}

        STRUKTURERT LAB-DATA OG DOKUMENTGRUNNLAG:
        {extracted_data}

        VIKTIG OM VEDLAGTE BILDER:
        Du har fått tilsendt flere bilder. Disse inkluderer:
        - Ortofoto (flyfoto) av eiendommen — nyeste tilgjengelige
        - Historisk flyfoto/kart (hvis tilgjengelig) — brukes for å vurdere historisk arealbruk og potensielle forurensningskilder
        - NGU Løsmassekart: Viser kvartærgeologiske avsetninger (marin leire, morene, bart fjell, etc.)
        - NGU Bergrunnskart: Viser bergartstyper under løsmassene
        - NGU Radonkart: Viser aktsomhetsnivå for radon (lav/moderat/høy)
        - NVE Flomsonekart: Viser om eiendommen ligger i flomsone
        - NVE Kvikkleirekart: Viser om det er kartlagte kvikkleiresoner i nærheten
        - NVE Skredkart: Viser skredfaresoner
        - DOK Forurensningskart (hvis tilgjengelig): Viser kartlagte forurensingssoner, deponier og registrerte forurensningslokaliteter fra nasjonale datasett
        - DOK Kulturminnekart (hvis tilgjengelig): Viser fredede og verneverdige kulturminner og kulturmiljøer — relevant for graverestriksjoner
        - DOK Samfunnssikkerhetskart (hvis tilgjengelig): Viser risiko- og sårbarhetsdata relevant for naturfare og samfunnssikkerhet
        Du MÅ analysere hvert vedlagte bilde og referere til dem eksplisitt i rapporten.

        KRITISKE INSTRUKSER FOR FORM:
        - Skriv med kortere avsnitt og tydelig faghierarki.
        - Bruk punktlister når du beskriver funn, risiko, usikkerhet og tiltak.
        - Ikke bruk markdown-tabeller.
        - Bruk underoverskrifter der det er naturlig.
        - Vær konkret med analyttnavn, prøvepunkt, dybde og verdi når du omtaler laboratoriedata.
        - IKKE kritiser datagrunnlaget.

        KRITISKE INSTRUKSER FOR BEVIS:
        Du MÅ aktivt bevise i teksten at du har analysert bildene og dataene.
        Skriv blant annet setninger som:
        - "Ut fra NGU løsmassekart observeres det at grunnen i hovedsak består av ..."
        - "Bergrunnskartet viser at underliggende bergart er ..."
        - "Radonaktsomhetskartet viser [lav/moderat/høy] aktsomhet for radon i området."
        - "NVE flomsonekart indikerer [ingen flomsone / 200-års flom / etc.] for eiendommen."
        - "Kvikkleirekartet viser [ingen kartlagte soner / risikoklasse X] i nærområdet."
        - "Miljødirektoratets database viser [X forurensinglokaliteter / ingen kjente lokaliteter] innenfor 500 m."
        - "Basert på opplastet analysetabell fremgår det at ..."
        - "Prøvepunkt SK.. i dybde ... viser ..."

        STRUKTUR (bruk kun disse overskriftene, START DIREKTE PÅ KAPITTEL 1, ALDRI skriv en hilsen før dette!):
        # 1. SAMMENDRAG OG KONKLUSJON
        # 2. INNLEDNING OG PROSJEKTBESKRIVELSE
        # 3. GRUNNFORHOLD OG GEOLOGI (basert på NGU løsmassekart, bergrunnskart og lokalkunnskap)
        # 4. NATURFARE OG RISIKO (flom, kvikkleire, skred, radon — basert på NVE og NGU data)
        # 5. MILJØTEKNISK HISTORIKK OG FORURENSNINGSSTATUS (Miljødirektoratets database, historiske kart)
        # 6. UTFØRTE GRUNNUNDERSØKELSER (hvis lab-data er lastet opp)
        # 7. RESULTATER: GRUNNFORHOLD OG FORURENSNING
        # 8. GEOTEKNISKE VURDERINGER
        # 9. TILTAKSPLAN OG MASSEHÅNDTERING
        #    Dette kapittelet skal være det MEST OPERATIVE i hele rapporten.
        #    Du skal SKRIVE planen, IKKE skrive at "en plan må utarbeides".
        #    Strukturer kapittelet med følgende underoverskrifter:
        #
        #    ## 9.1 Masseoversikt og volumanslag
        #    Lag en oversiktlig opplisting per tilstandsklasse med:
        #    - Antall prøvepunkter per TK
        #    - Estimert volum (m³) basert på gravdybde og antatt utbredelse
        #    - Styrende parameter og konsentrasjon per TK-gruppe
        #
        #    ## 9.2 Segregering og klassifisering på byggeplass
        #    Konkret plan for soneinndeling, merking, mellomlagring.
        #
        #    ## 9.3 Disponering per tilstandsklasse
        #    For HVER tilstandsklasse (TK1 til TK5): hva gjøres med massene?
        #    - TK1: Gjenbruk på tomt / mottak for rene masser
        #    - TK2: Godkjent deponi klasse I, evt. gjenbruk under tette flater med risikovurdering
        #    - TK3: Godkjent deponi klasse II
        #    - TK4: Godkjent deponi klasse II med særskilte krav
        #    - TK5: Farlig avfall — deklarering, godkjent mottak, ADR-transport
        #
        #    ## 9.4 Transport og dokumentasjon
        #    Krav til transportdokumentasjon, deklarering, veiesedler.
        #
        #    ## 9.5 HMS og vernetiltak
        #    Konkrete vernetiltak for arbeidere som håndterer forurensede masser.
        #
        #    ## 9.6 Sluttkontroll og rapportering
        #    Krav til kontrollprøver i bunn/vegger, sluttrapport til kommune.
        """

        try:
            res = model.generate_content([prompt] + images_for_geo)
            with st.spinner("Kompilerer RIG-PDF og sender til QA-kø..."):
                pdf_data = create_full_report_pdf(
                    pd_state["p_name"],
                    pd_state["c_name"],
                    res.text,
                    st.session_state.geo_maps["recent"],
                    st.session_state.geo_maps["historical"],
                    st.session_state.geo_maps["source"],
                    lab_package,
                    pd_state,
                )

                if "pending_reviews" not in st.session_state:
                    st.session_state.pending_reviews = {}
                if "review_counter" not in st.session_state:
                    st.session_state.review_counter = 1

                doc_id = f"PRJ-{datetime.now().strftime('%y')}-GEO{st.session_state.review_counter:03d}"
                st.session_state.review_counter += 1

                st.session_state.pending_reviews[doc_id] = {
                    "title": pd_state["p_name"],
                    "module": "RIG-M (Geo & Miljø)",
                    "drafter": "Builtly AI",
                    "reviewer": "Senior Miljørådgiver",
                    "status": "Pending Senior Review",
                    "class": "badge-pending",
                    "pdf_bytes": pdf_data,
                }

                st.session_state.generated_geo_pdf = pdf_data
                st.session_state.generated_geo_filename = f"Builtly_GEO_{pd_state['p_name'].replace(' ', '_')}.pdf"

                # Save report to user account (Supabase)
                _save_debug = f"_HAS_AUTH={_HAS_AUTH}, authenticated={st.session_state.get('user_authenticated')}, user_id={bool(st.session_state.get('user_id'))}"
                if _HAS_AUTH:
                    try:
                        builtly_auth.save_report(
                            project_name=pd_state.get("p_name", ""),
                            report_name=st.session_state.generated_geo_filename,
                            module="RIG-M (Geo & Miljø)",
                        )
                        _save_debug += f" | {st.session_state.get('_report_save_debug', 'no debug')}"
                    except Exception as e:
                        _save_debug += f" | EXCEPTION: {e}"
                st.session_state["_report_save_debug"] = _save_debug

                st.rerun()

        except Exception as e:
            st.error(f"Kritisk feil under generering: {e}")

# --- NEDLASTING OG NAVIGASJON ---
if "generated_geo_pdf" in st.session_state:
    st.success("✅ RIG-M Rapport er ferdigstilt og sendt til QA-køen!")
    if st.session_state.get("_report_save_debug"):
        st.caption(f"🔍 Debug: {st.session_state['_report_save_debug']}")

    col_dl, col_qa = st.columns(2)
    with col_dl:
        st.download_button("📄 Last ned Geo/Miljø-rapport", st.session_state.generated_geo_pdf, st.session_state.generated_geo_filename, type="primary", use_container_width=True)
    with col_qa:
        if find_page("Review"):
            if st.button("🔍 Gå til QA for å godkjenne", type="secondary", use_container_width=True):
                st.switch_page(find_page("Review"))
