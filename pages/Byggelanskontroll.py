# -*- coding: utf-8 -*-
"""
Builtly | Byggelånskontroll v3
Oppgradert modul med automatisk dokumentlesing og pre-utfylling.
Self-contained Streamlit module – no external builtly_* dependencies.
"""
from __future__ import annotations

import base64, io, json, os, re, textwrap
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st

try:
    from fpdf import FPDF
except ImportError:
    FPDF = None

try:
    from PIL import Image
except ImportError:
    Image = None

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import fitz  # PyMuPDF (fallback)
except ImportError:
    fitz = None

try:
    from openai import OpenAI
except Exception:
    OpenAI = None

try:
    import google.generativeai as genai
except Exception:
    genai = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Builtly | Byggelånskontroll", layout="wide", initial_sidebar_state="collapsed")


# ────────────────────────────────────────────────────────────────
# HELPERS
# ────────────────────────────────────────────────────────────────
def render_html(html: str) -> None:
    st.markdown(html.replace("\n", " "), unsafe_allow_html=True)


def logo_data_uri() -> str:
    for candidate in ["logo-white.png", "logo.png"]:
        if os.path.exists(candidate):
            suffix = Path(candidate).suffix.lower().replace(".", "")
            mime = f"image/{'jpeg' if suffix in ('jpg','jpeg') else suffix}"
            with open(candidate, "rb") as f:
                return f"data:{mime};base64,{base64.b64encode(f.read()).decode()}"
    return ""


def render_hero(eyebrow, title, subtitle, pills, badge):
    pills_html = "".join(f'<span class="hero-pill">{p}</span>' for p in pills)
    render_html(f"""
    <div class="hero-card">
        <div class="hero-eyebrow">{eyebrow}</div>
        <div class="hero-title">{title}</div>
        <div class="hero-subtitle">{subtitle}</div>
        <div class="hero-pills">{pills_html}</div>
        <div class="hero-badge">{badge}</div>
    </div>""")


def render_section(title, desc, badge):
    render_html(f"""
    <div class="section-header">
        <span class="section-badge">{badge}</span>
        <h3>{title}</h3>
        <p>{desc}</p>
    </div>""")


def render_panel(title, desc, bullets, tone="blue", badge=""):
    color_map = {"blue": ("#38bdf8", "rgba(56,194,201,0.06)", "rgba(56,194,201,0.18)"),
                 "gold": ("#f59e0b", "rgba(245,158,11,0.06)", "rgba(245,158,11,0.18)"),
                 "green": ("#22c55e", "rgba(34,197,94,0.06)", "rgba(34,197,94,0.18)"),
                 "red": ("#ef4444", "rgba(239,68,68,0.06)", "rgba(239,68,68,0.18)")}
    accent, bg, border = color_map.get(tone, color_map["blue"])
    badge_html = f'<span style="display:inline-block;background:{bg};border:1px solid {border};border-radius:6px;padding:1px 8px;font-size:0.7rem;font-weight:700;color:{accent};text-transform:uppercase;letter-spacing:0.06em;margin-bottom:6px;">{badge}</span>' if badge else ""
    bullets_html = "".join(f'<li style="color:#c8d3df;margin-bottom:6px;font-size:0.88rem;line-height:1.5;">{b}</li>' for b in bullets)
    render_html(f"""
    <div class="panel-box" style="background:{bg};border:1px solid {border};border-radius:14px;padding:1.3rem 1.5rem;margin-bottom:1rem;">
        {badge_html}
        <div style="font-weight:700;font-size:0.98rem;color:#f5f7fb;margin-bottom:4px;">{title}</div>
        <div style="font-size:0.85rem;color:#9fb0c3;margin-bottom:10px;line-height:1.5;">{desc}</div>
        <ul style="margin:0;padding-left:1.2rem;">{bullets_html}</ul>
    </div>""")


def render_metric_cards(metrics):
    cards = ""
    for val, label, desc in metrics:
        cards += f"""<div class="metric-card">
            <div class="mc-value">{val}</div>
            <div class="mc-label">{label}</div>
            <div class="mc-desc">{desc}</div>
        </div>"""
    render_html(f'<div class="metric-row">{cards}</div>')


def safe_get(obj, key, default=""):
    if isinstance(obj, dict):
        return obj.get(key, default)
    return default


def fmt_nok(amount: float) -> str:
    """Format number as NOK with thousand separators."""
    if amount >= 1_000_000:
        return f"{amount/1_000_000:,.1f} MNOK".replace(",", " ")
    elif amount >= 1_000:
        return f"{amount:,.0f} kr".replace(",", " ")
    return f"{amount:,.2f} kr".replace(",", " ")


# ────────────────────────────────────────────────────────────────
# DOCUMENT AUTO-EXTRACTION ENGINE
# ────────────────────────────────────────────────────────────────

def extract_pdf_text(raw_bytes: bytes) -> str:
    """Extract text from PDF bytes using pdfplumber (preferred) or PyMuPDF (fallback)."""
    # Try pdfplumber first (handles most PDFs well)
    if pdfplumber:
        try:
            with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
                text = "\n".join(p.extract_text() or "" for p in pdf.pages)
                if text.strip():
                    return text
        except Exception:
            pass
    # Fallback to PyMuPDF
    if fitz:
        try:
            doc = fitz.open(stream=raw_bytes, filetype="pdf")
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            if text.strip():
                return text
        except Exception:
            pass
    return ""


def extract_excel_data(raw_bytes: bytes) -> Dict[str, pd.DataFrame]:
    """Extract all sheets from Excel file as DataFrames."""
    sheets = {}
    try:
        xls = pd.ExcelFile(io.BytesIO(raw_bytes))
        for name in xls.sheet_names:
            sheets[name] = pd.read_excel(xls, sheet_name=name)
    except Exception:
        pass
    return sheets


def extract_excel_raw(raw_bytes: bytes) -> Dict[str, list]:
    """Extract raw cell data from Excel using openpyxl for better structure detection."""
    if not openpyxl:
        return {}
    result = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=False):
                row_data = {c.coordinate: c.value for c in row if c.value is not None}
                if row_data:
                    rows.append(row_data)
            result[sheet_name] = rows
        wb.close()
    except Exception:
        pass
    return result


def classify_document(filename: str, text: str) -> str:
    """Classify a document based on filename and content."""
    name_lower = filename.lower()
    text_lower = text.lower() if text else ""

    # Kalkyle / prosjektkalkyle
    if any(k in name_lower for k in ["kalkyle", "kalkyl"]) or "prosjektkalkyle" in text_lower:
        return "kalkyle"

    # Sjekkliste
    if "sjekkliste" in name_lower or "sjekkliste byggelånskontroll" in text_lower:
        return "sjekkliste"

    # Byggelånsrapport
    if "byggelånsrapport" in name_lower or "byggelansrapport" in name_lower or "byggelånsrapport" in text_lower:
        return "byggelaansrapport"

    # Opptrekk / trekkanmodning
    if any(k in name_lower for k in ["opptrekk", "trekk", "trekkanmodning"]):
        return "opptrekk"

    # Faktura
    if "faktura" in name_lower or "faktura" in text_lower:
        return "faktura"

    # DokumentID (often invoices)
    if "dokumentid" in name_lower:
        return "faktura"

    # E-post / korrespondanse
    if "e-post" in name_lower or "epost" in name_lower:
        return "epost"

    # Entreprisekontrakt
    if any(k in name_lower for k in ["entreprise", "kontrakt"]):
        return "kontrakt"

    # Fremdriftsplan
    if "fremdrift" in name_lower:
        return "fremdrift"

    # Betalingsplan
    if "betaling" in name_lower:
        return "betalingsplan"

    return "ukjent"


def parse_number(text: str) -> Optional[float]:
    """Parse Norwegian formatted number from text."""
    if not text:
        return None
    # Remove spaces and replace comma with dot
    cleaned = re.sub(r'\s+', '', str(text))
    cleaned = cleaned.replace(',', '.')
    # Remove trailing dots
    cleaned = cleaned.rstrip('.')
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def extract_from_kalkyle(text: str) -> dict:
    """Extract project data from a prosjektkalkyle PDF."""
    data = {}

    # Prosjektnummer
    m = re.search(r'Prosjektnummer[:\s]*(\d+)', text)
    if m:
        data['prosjektnummer'] = m.group(1)

    # BT / byggetrinn
    m = re.search(r'BT\s*(\d+)', text)
    if m:
        data['byggetrinn'] = m.group(1)

    # BRA (handles "BRA 7 644" and "BRA7 644")
    m = re.search(r'BRA\s+(\d[\d\s]*\d)', text)
    if m:
        data['bra'] = parse_number(m.group(1))

    # BRAs (handles dashes and spaces: "BRAS - 6 201-")
    m = re.search(r'BRAS?\s*[-–]?\s*(\d[\d\s]*\d)', text, re.IGNORECASE)
    if m:
        data['bras'] = parse_number(m.group(1))

    # Antall enheter bolig (handles "Antall enheter bolig - 130-")
    m = re.search(r'Antall\s+enheter\s+bolig\s*[-–]?\s*(\d+)', text)
    if m:
        data['antall_enheter'] = int(m.group(1))

    # Antall garasje/parkering (handles dashes)
    m = re.search(r'Antall\s+garasje/parkering\s*[-–]?\s*(\d+)', text)
    if m:
        data['antall_parkering'] = int(m.group(1))

    # Salgsgrad
    m = re.search(r'Salgsgrad\s+(\d+)%', text)
    if m:
        data['salgsgrad'] = int(m.group(1))

    # Produksjonsgrad
    m = re.search(r'Produksjonsgrad\s+(\d+)%', text)
    if m:
        data['produksjonsgrad'] = int(m.group(1))

    # Fullføringsgrad
    m = re.search(r'Fullføringsgrad\s+(\d+)%', text)
    if m:
        data['fullfoeringsgrad'] = int(m.group(1))

    # Eierandel (handles "50,00%")
    m = re.search(r'Eierandel\s+([\d,]+)%', text)
    if m:
        data['eierandel'] = parse_number(m.group(1).replace(',', '.'))

    # Resultat prosent (handles "11,7 %")
    m = re.search(r'Resultat\s*%\s*[-–]?\s*før\s*skatt\s+([\d,]+)\s*%', text)
    if m:
        data['resultat_prosent'] = parse_number(m.group(1))

    # Inntekter totalt - line format: "Inntekter totalt 0 573 878 624 330 035 800 ..."
    m = re.search(r'[Ii]nntekter\s+totalt\s+\d+\s+([\d\s]+?)(?:\s+\d[\d\s]+){3}', text)
    if m:
        val = parse_number(m.group(1))
        if val and val > 1000:
            data['total_inntekter'] = val

    # Kostnader totalt - line format: "Kostnader totalt 0 506 451 706 155 490 410 ..."
    m = re.search(r'[Kk]ostnader\s+totalt\s+\d+\s+([\d\s]+?)(?:\s+\d[\d\s]+){3}', text)
    if m:
        val = parse_number(m.group(1))
        if val and val > 1000:
            data['totalbudsjett'] = val

    # Fallback: extract from "Kostnader totalt" line with position-based parsing
    if 'totalbudsjett' not in data:
        for line in text.split('\n'):
            if 'Kostnader totalt' in line or 'ostnader totalt' in line:
                # Remove label and leading zero, then take first 11-char group (NNN NNN NNN)
                after = re.sub(r'.*?totalt\s+\d\s+', '', line).strip()
                if len(after) >= 11:
                    budget_str = after[:11].strip()
                    try:
                        data['totalbudsjett'] = int(budget_str.replace(' ', ''))
                    except ValueError:
                        pass
                    # Second column: påløpt
                    if len(after) >= 23:
                        paloept_str = after[12:23].strip()
                        try:
                            data['paloept_kostnader'] = int(paloept_str.replace(' ', ''))
                        except ValueError:
                            pass
                break

    if 'total_inntekter' not in data:
        for line in text.split('\n'):
            if 'Inntekter totalt' in line or 'nntekter totalt' in line:
                after = re.sub(r'.*?totalt\s+\d\s+', '', line).strip()
                if len(after) >= 11:
                    innt_str = after[:11].strip()
                    try:
                        data['total_inntekter'] = int(innt_str.replace(' ', ''))
                    except ValueError:
                        pass
                break

    # Resultat - før skatt (number) - use position-based extraction
    for line in text.split('\n'):
        if re.search(r'Resultat\s*[-–]\s*før\s*skatt', line) and '%' not in line:
            after = re.sub(r'Resultat\s*[-–]\s*før\s*skatt\s*', '', line).strip()
            if len(after) >= 10:
                result_str = after[:12].strip()
                try:
                    data['resultat'] = int(result_str.replace(' ', '').replace('−', '-'))
                except ValueError:
                    pass
            break

    return data


def extract_from_sjekkliste(text: str) -> dict:
    """Extract data from sjekkliste for byggelånskontroll."""
    data = {}

    # Byggherre
    m = re.search(r'Byggherre[:\s]+([^\n]+)', text)
    if m:
        data['byggherre'] = m.group(1).strip()

    # Prosjektnavn
    m = re.search(r'Prosjektnavn[:\s]+([^\n(]+)', text)
    if m:
        data['prosjektnavn'] = m.group(1).strip()

    # Byggelånskontrollør (handles garbled text like "ntrollør: Stein Johnsen")
    m = re.search(r'ntrollør:\s*([A-ZÆØÅ][a-zæøå]+\s+[A-ZÆØÅ][a-zæøå]+)', text)
    if m:
        data['kontrolloer'] = m.group(1).strip()

    # Extract checklist items (Ja/Nei status)
    checklist_items = {}
    lines = text.split('\n')
    for line in lines:
        if re.search(r'\bX\b', line):
            item_name = line.split('X')[0].strip()
            if item_name and len(item_name) > 3:
                checklist_items[item_name] = True
    if checklist_items:
        data['sjekkliste_items'] = checklist_items

    # Entreprisekontrakter
    if 'Totalentreprise' in text:
        data['entrepriseform'] = 'Totalentreprise'
        m = re.search(r'Totalentrepriseavtale\s+(\w[\w\s]+)', text)
        if m:
            data['totalentreprenoer'] = m.group(1).strip()

    return data


def extract_from_byggelaansrapport(text: str) -> dict:
    """Extract data from byggelånsrapport."""
    data = {}

    # Rapport nr - handles "Rapport nr. Rev.nr.\n... 02 02"
    m = re.search(r'Rapport\s+nr\.?\s*(?:Rev\.?\s*nr\.?)?\s*\n.*?(\d+)\s+(\d+)', text)
    if m:
        data['rapport_nr'] = int(m.group(2))
    else:
        m = re.search(r'Rapport\s*nr\.?\s*(\d+)', text)
        if m:
            data['rapport_nr'] = int(m.group(1))

    # Dato
    m = re.search(r'(?:Dato|dato)\s+([\d.]+\.\d{4})', text)
    if m:
        data['rapport_dato'] = m.group(1)

    # Forfatter / kontrollør
    m = re.search(r'(?:Forfatter|Stein\s+Johnsen)', text)
    if m:
        # Look for name after Forfatter or just find "Stein Johnsen"
        n = re.search(r'Stein\s+Johnsen', text)
        if n:
            data['kontrolloer'] = 'Stein Johnsen'

    # Anbefalt opptrekk - handles multiline "opptrekk på byggelånskrediten med\nkr 9 542 229,-"
    m = re.search(r'opptrekk.*?kr\s*([\d\s]+)', text, re.IGNORECASE | re.DOTALL)
    if m:
        raw = m.group(1).strip()
        val = int(raw.replace(' ', '')) if raw.replace(' ', '').isdigit() else None
        if val and val > 1000:
            data['anbefalt_opptrekk'] = val

    # Verdi av nedlagte arbeider
    m = re.search(r'pålydende\s+kr\s*([\d\s]+)', text)
    if m:
        raw = m.group(1).strip()
        val = int(raw.replace(' ', '')) if raw.replace(' ', '').isdigit() else None
        if val and val > 1000:
            data['verdi_nedlagt'] = val

    # Inngåtte entrepriser - "Spennarmering AS (inkl...) kr 342 486 750,00"
    entrepriser = []
    for m in re.finditer(r'([\w\s]+(?:AS|as))\s+\(.*?\)\s*kr\s*([\d\s.,]+)', text):
        name = m.group(1).strip()
        raw = m.group(2).strip().rstrip('.,- ')
        val = parse_number(raw.replace('.', '').replace(',', '.'))
        if not val:
            val = parse_number(raw.replace(' ', ''))
        if val:
            entrepriser.append({'firma': name, 'beloep': val})
    if entrepriser:
        data['entrepriser'] = entrepriser

    # Forsikringer
    forsikringer = []
    for m in re.finditer(r'(\w[\w\s]+(?:AS|as))\s+([\w\s-]*forsikring)\s+([\d.]+\.\d{4})', text):
        forsikringer.append({
            'firma': m.group(1).strip(),
            'type': m.group(2).strip(),
            'gyldig_til': m.group(3)
        })
    if forsikringer:
        data['forsikringer'] = forsikringer

    # Status fremdrift bullet points
    fremdrift_items = []
    in_status = False
    for line in text.split('\n'):
        if 'Status fremdrift' in line:
            in_status = True
            continue
        if in_status and line.strip().startswith('•'):
            fremdrift_items.append(line.strip().lstrip('• '))
        elif in_status and not line.strip().startswith('•') and line.strip():
            if 'Betalingsplan' in line or 'Konklusjon' in line or 'Forskudd' in line:
                in_status = False
    if fremdrift_items:
        data['fremdrift_status'] = fremdrift_items

    # Manglende dokumentasjon
    mangler = []
    text_lower = text.lower()
    if 'etterlyst' in text_lower and 'ikke mottatt' in text_lower:
        if 'betalingsplan' in text_lower:
            mangler.append('Betalingsplan for totalentreprisen - etterlyst men ikke mottatt')
        if 'garanti' in text_lower and 'ikke mottatt' in text_lower:
            mangler.append('Garantistillelse fra entreprenør - etterlyst men ikke mottatt')
    # Check for "Firma AS   Ikke mottatt" on clean lines
    for line in text.split('\n'):
        line_stripped = line.strip()
        if re.match(r'^[A-ZÆØÅ][\w\s]+AS\s+[Ii]kke\s+mottatt', line_stripped):
            m_firm = re.match(r'^([A-ZÆØÅ][\w\s]+AS)', line_stripped)
            if m_firm:
                mangler.append(f'Dokumentasjon fra {m_firm.group(1).strip()} - ikke mottatt')
    if mangler:
        data['manglende_dokumenter'] = list(set(mangler))

    return data


def extract_from_opptrekk_excel(raw_bytes: bytes) -> dict:
    """Extract draw request data from opptrekk Excel file."""
    data = {}

    raw_data = extract_excel_raw(raw_bytes)
    sheets = extract_excel_data(raw_bytes)

    # Parse "Oversikt byggelån" sheet
    if 'Oversikt byggelån' in raw_data:
        for row in raw_data['Oversikt byggelån']:
            for coord, val in row.items():
                if isinstance(val, str):
                    val_clean = val.strip().lower()
                    if 'ramme' in val_clean and 'total' in val_clean:
                        # Next cell should have the value
                        col_letter = coord[0]
                        row_num = coord[1:]
                        next_coord = f"B{row_num}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['laaneramme'] = row[next_coord]
                    elif 'åpningstrekk' in val_clean or 'trekk 1' in val_clean:
                        next_coord = f"B{coord[1:]}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['aapningstrekk'] = row[next_coord]
                    elif 'trekkanmodning' in val_clean:
                        next_coord = f"B{coord[1:]}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['dette_trekk'] = row[next_coord]
                            # Extract trekk number
                            nm = re.search(r'(\d+)', val)
                            if nm:
                                data['trekk_nr'] = int(nm.group(1))
                    elif 'totalt' in val_clean and 'rest' not in val_clean:
                        next_coord = f"B{coord[1:]}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['totalt_trukket'] = row[next_coord]
                    elif 'rest byggelån' in val_clean:
                        next_coord = f"B{coord[1:]}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['rest_byggelaan'] = row[next_coord]
                    elif 'ønsket opptrekk' in val_clean:
                        next_coord = f"B{coord[1:]}"
                        if next_coord in row and isinstance(row[next_coord], (int, float)):
                            data['oensket_opptrekk'] = row[next_coord]
                        # Extract date
                        dm = re.search(r'(\d{2}\.\d{2}\.\d{4})', val)
                        if dm:
                            data['opptrekk_dato'] = dm.group(1)

    # Parse transaction sheets for line items
    for sheet_name in ['Prosjekt BT3', 'Prosjekt infra']:
        if sheet_name in sheets:
            df = sheets[sheet_name]
            if 'Beløp' in df.columns or 'L' in df.columns:
                beloep_col = 'Beløp' if 'Beløp' in df.columns else 'L'
                total = df[beloep_col].sum() if beloep_col in df.columns else 0
                data[f'sum_{sheet_name.lower().replace(" ", "_")}'] = total

    # Extract individual line items for verification
    line_items = []
    for sheet_name in ['Prosjekt BT3', 'Prosjekt infra']:
        if sheet_name in sheets:
            df = sheets[sheet_name]
            for _, row in df.iterrows():
                item = {}
                if 'Kundenavn' in df.columns:
                    item['leverandoer'] = row.get('Kundenavn', '')
                elif 'Q' in df.columns:
                    item['leverandoer'] = row.get('Q', '')
                if 'Beløp' in df.columns:
                    item['beloep'] = row.get('Beløp', 0)
                elif 'L' in df.columns:
                    item['beloep'] = row.get('L', 0)
                if 'Bilagsdato' in df.columns:
                    item['dato'] = str(row.get('Bilagsdato', ''))
                elif 'J' in df.columns:
                    item['dato'] = str(row.get('J', ''))
                if 'Produktkode Navn' in df.columns:
                    item['kategori'] = row.get('Produktkode Navn', '')
                elif 'AJ' in df.columns:
                    item['kategori'] = row.get('AJ', '')
                if 'Fakturanr.' in df.columns:
                    item['fakturanr'] = str(row.get('Fakturanr.', ''))
                elif 'D' in df.columns:
                    item['fakturanr'] = str(row.get('D', ''))
                if 'DokumentID' in df.columns:
                    item['dokument_id'] = str(row.get('DokumentID', ''))
                elif 'C' in df.columns:
                    item['dokument_id'] = str(row.get('C', ''))
                if any(v for v in item.values()):
                    line_items.append(item)
    if line_items:
        data['bilag'] = line_items

    # Parse infrastructure allocation
    if 'Fordeling infrastruktur' in raw_data:
        for row in raw_data['Fordeling infrastruktur']:
            for coord, val in row.items():
                if isinstance(val, str) and 'bt3' in val.lower():
                    # Look for the amount in column D
                    d_coord = f"D{coord[1:]}"
                    if d_coord in row and isinstance(row[d_coord], (int, float)):
                        data['infra_andel_bt3'] = row[d_coord]

    return data


def extract_from_faktura(text: str) -> dict:
    """Extract data from invoice PDF. Falls back to minimal data if image-based."""
    data = {}

    # If no text was extracted (image-based PDF), flag it
    if not text or len(text.strip()) < 10:
        data['_bildebasert'] = True
        data['_kommentar'] = 'Bildebasert PDF — krever AI-analyse eller OCR for fullstendig lesing'
        return data

    # Fakturanr
    m = re.search(r'Fakturanr\.?\s*:?\s*(\d+)', text)
    if m:
        data['fakturanr'] = m.group(1)

    # Fra (leverandør)
    m = re.search(r'Fra\s+(\w[\w\s]+(?:AS|as))', text)
    if m:
        data['leverandoer'] = m.group(1).strip()

    # Til (kunde)
    m = re.search(r'Til\s+(\w[\w\s]+(?:AS|as))', text)
    if m:
        data['kunde'] = m.group(1).strip()

    # Totalt NOK / Å betale
    for pattern_name, pattern in [
        ("total_inkl_mva", r'Totalt\s+NOK\s+([\d\s.,]+)'),
        ("total_inkl_mva", r'[Åå]\s+betale\s+([\d\s.,]+)'),
    ]:
        m = re.search(pattern, text)
        if m and pattern_name not in data:
            raw = m.group(1).strip().rstrip('.,- ')
            val = parse_number(raw.replace('.', '').replace(',', '.'))
            if val:
                data[pattern_name] = val

    # Beløp eks MVA
    m = re.search(r'Totalt,?\s*uten\s+MVA\s+([\d\s.,]+)', text)
    if m:
        raw = m.group(1).strip().rstrip('.,- ')
        val = parse_number(raw.replace('.', '').replace(',', '.'))
        if val:
            data['total_eks_mva'] = val

    # Sum (alternative)
    m = re.search(r'Sum\s+([\d\s.,]+)', text)
    if m and 'total_eks_mva' not in data:
        raw = m.group(1).strip().rstrip('.,- ')
        val = parse_number(raw.replace('.', '').replace(',', '.'))
        if val:
            data['total_eks_mva'] = val

    # Forfallsdato
    m = re.search(r'[Ff]orfallsdato\s+([\d.]+)', text)
    if m:
        data['forfallsdato'] = m.group(1)

    # Dato
    m = re.search(r'Dato\s+([\d.]+\.\d{4})', text)
    if m:
        data['faktura_dato'] = m.group(1)

    # KID
    m = re.search(r'KID[:\s]+([\d]+)', text)
    if m:
        data['kid'] = m.group(1)

    # Kontonr / Bankkonto
    m = re.search(r'(?:Kontonr|Bankkonto)[.:\s]+([\d]+)', text)
    if m:
        data['kontonr'] = m.group(1)

    # Innestående (7.5% etc.)
    m = re.search(r'(\d+[.,]\d+)%\s+innestående', text, re.IGNORECASE)
    if m:
        data['innestaaende_prosent'] = parse_number(m.group(1).replace(',', '.'))

    return data


def auto_extract_all(files: list) -> dict:
    """
    Master extraction function. Processes all uploaded files and returns
    a consolidated project data dictionary with confidence markers.
    """
    extracted = {
        'documents': [],      # List of classified documents
        'project': {},        # Consolidated project data
        'opptrekk': {},       # Draw request specifics
        'fakturaer': [],      # List of invoices
        'sjekkliste': {},     # Checklist data
        'rapport': {},        # Byggelånsrapport data
        'kalkyle': {},        # Budget/kalkyle data
        'bilag': [],          # Line items from opptrekk
        'confidence': {},     # Confidence score per field
        'mangler': [],        # Missing items identified
    }

    for f in files:
        raw = f.read()
        f.seek(0)
        name = f.name
        name_lower = name.lower()

        text = ""
        doc_data = {}

        # Extract text/data based on file type
        if name_lower.endswith('.pdf'):
            text = extract_pdf_text(raw)
            doc_type = classify_document(name, text)

            if doc_type == 'kalkyle':
                doc_data = extract_from_kalkyle(text)
                extracted['kalkyle'].update(doc_data)
            elif doc_type == 'sjekkliste':
                doc_data = extract_from_sjekkliste(text)
                extracted['sjekkliste'].update(doc_data)
            elif doc_type == 'byggelaansrapport':
                doc_data = extract_from_byggelaansrapport(text)
                extracted['rapport'].update(doc_data)
            elif doc_type == 'faktura':
                doc_data = extract_from_faktura(text)
                extracted['fakturaer'].append(doc_data)

        elif name_lower.endswith(('.xlsx', '.xls', '.xlsm')):
            doc_type = classify_document(name, "")
            if doc_type == 'opptrekk' or 'opptrekk' in name_lower or 'trekk' in name_lower:
                doc_data = extract_from_opptrekk_excel(raw)
                extracted['opptrekk'].update(doc_data)
                if 'bilag' in doc_data:
                    extracted['bilag'] = doc_data['bilag']
            else:
                # Try generic Excel extraction
                sheets = extract_excel_data(raw)
                doc_type = 'excel'
                doc_data = {'sheets': list(sheets.keys())}

        elif name_lower.endswith(('.csv', '.txt', '.md')):
            try:
                text = raw.decode('utf-8', errors='replace')
            except Exception:
                pass
            doc_type = 'tekst'

        elif name_lower.endswith(('.jpg', '.jpeg', '.png', '.gif')):
            doc_type = 'bilde'

        else:
            doc_type = 'ukjent'

        extracted['documents'].append({
            'filename': name,
            'type': doc_type,
            'extracted_fields': list(doc_data.keys()) if isinstance(doc_data, dict) else [],
            'text_length': len(text),
        })

    # ── Consolidate project data ──
    proj = extracted['project']

    # From sjekkliste
    if extracted['sjekkliste']:
        sj = extracted['sjekkliste']
        if 'byggherre' in sj:
            proj['utbygger'] = sj['byggherre']
            extracted['confidence']['utbygger'] = 'høy'
        if 'prosjektnavn' in sj:
            proj['prosjektnavn'] = sj['prosjektnavn']
            extracted['confidence']['prosjektnavn'] = 'høy'
        if 'kontrolloer' in sj:
            proj['kontrolloer'] = sj['kontrolloer']
            extracted['confidence']['kontrolloer'] = 'høy'
        if 'entrepriseform' in sj:
            proj['entrepriseform'] = sj['entrepriseform']
            extracted['confidence']['entrepriseform'] = 'høy'
        if 'totalentreprenoer' in sj:
            proj['totalentreprenoer'] = sj['totalentreprenoer']
            extracted['confidence']['totalentreprenoer'] = 'høy'

    # From kalkyle
    if extracted['kalkyle']:
        kal = extracted['kalkyle']
        if 'totalbudsjett' in kal:
            proj['totalbudsjett'] = kal['totalbudsjett']
            extracted['confidence']['totalbudsjett'] = 'høy'
        if 'total_inntekter' in kal:
            proj['total_inntekter'] = kal['total_inntekter']
            extracted['confidence']['total_inntekter'] = 'høy'
        if 'salgsgrad' in kal:
            proj['salgsgrad'] = kal['salgsgrad']
            extracted['confidence']['salgsgrad'] = 'høy'
        if 'produksjonsgrad' in kal:
            proj['produksjonsgrad'] = kal['produksjonsgrad']
            extracted['confidence']['produksjonsgrad'] = 'høy'
        if 'fullfoeringsgrad' in kal:
            proj['fullfoeringsgrad'] = kal['fullfoeringsgrad']
            extracted['confidence']['fullfoeringsgrad'] = 'høy'
        if 'antall_enheter' in kal:
            proj['antall_enheter'] = kal['antall_enheter']
            extracted['confidence']['antall_enheter'] = 'høy'
        if 'eierandel' in kal:
            proj['eierandel'] = kal['eierandel']
            extracted['confidence']['eierandel'] = 'høy'
        if 'resultat_prosent' in kal:
            proj['resultat_prosent'] = kal['resultat_prosent']
            extracted['confidence']['resultat_prosent'] = 'høy'
        if 'prosjektnummer' in kal:
            proj['prosjektnummer'] = kal['prosjektnummer']
            extracted['confidence']['prosjektnummer'] = 'høy'

    # From opptrekk
    if extracted['opptrekk']:
        opp = extracted['opptrekk']
        if 'laaneramme' in opp:
            proj['byggelaan'] = opp['laaneramme']
            extracted['confidence']['byggelaan'] = 'høy'
        if 'dette_trekk' in opp or 'oensket_opptrekk' in opp:
            proj['forespurt_trekk'] = opp.get('oensket_opptrekk', opp.get('dette_trekk', 0))
            extracted['confidence']['forespurt_trekk'] = 'høy'
        if 'totalt_trukket' in opp:
            proj['tidligere_utbetalt'] = opp.get('aapningstrekk', opp['totalt_trukket'] - opp.get('dette_trekk', 0))
            extracted['confidence']['tidligere_utbetalt'] = 'høy'
        if 'trekk_nr' in opp:
            proj['trekk_nr'] = opp['trekk_nr']
            extracted['confidence']['trekk_nr'] = 'høy'
        if 'opptrekk_dato' in opp:
            proj['opptrekk_dato'] = opp['opptrekk_dato']
            extracted['confidence']['opptrekk_dato'] = 'høy'

    # From byggelånsrapport
    if extracted['rapport']:
        rap = extracted['rapport']
        if 'anbefalt_opptrekk' in rap:
            proj['anbefalt_opptrekk'] = rap['anbefalt_opptrekk']
            extracted['confidence']['anbefalt_opptrekk'] = 'høy'
        if 'rapport_nr' in rap:
            proj['rapport_nr'] = rap['rapport_nr']
            extracted['confidence']['rapport_nr'] = 'høy'
        if 'fremdrift_status' in rap:
            proj['fremdrift_status'] = rap['fremdrift_status']
        if 'manglende_dokumenter' in rap:
            extracted['mangler'].extend(rap['manglende_dokumenter'])
        if 'entrepriser' in rap:
            proj['entrepriser'] = rap['entrepriser']
        # Kontrollør fallback from rapport
        if 'kontrolloer' not in proj and 'kontrolloer' in rap:
            proj['kontrolloer'] = rap['kontrolloer']
            extracted['confidence']['kontrolloer'] = 'middels'

    return extracted


# ────────────────────────────────────────────────────────────────
# AI ENGINE
# ────────────────────────────────────────────────────────────────
def get_ai_client():
    """Returns (client_type, client) – 'openai'|'gemini'|None."""
    oai_key = os.environ.get("OPENAI_API_KEY", "")
    gem_key = os.environ.get("GOOGLE_API_KEY", "") or os.environ.get("GEMINI_API_KEY", "")
    if oai_key and OpenAI:
        return "openai", OpenAI(api_key=oai_key)
    if gem_key and genai:
        genai.configure(api_key=gem_key)
        return "gemini", genai.GenerativeModel("gemini-2.0-flash")
    return None, None


def extract_text_from_uploads(files) -> str:
    """Extracts text from uploaded PDFs/text files."""
    all_text = []
    for f in files:
        raw = f.read()
        f.seek(0)
        name = f.name.lower()
        if name.endswith(".pdf"):
            # Use pdfplumber first
            if pdfplumber:
                try:
                    with pdfplumber.open(io.BytesIO(raw)) as pdf:
                        for page in pdf.pages:
                            t = page.extract_text()
                            if t:
                                all_text.append(t)
                except Exception:
                    pass
            elif fitz:
                try:
                    doc = fitz.open(stream=raw, filetype="pdf")
                    for page in doc:
                        all_text.append(page.get_text())
                    doc.close()
                except Exception:
                    pass
        elif name.endswith((".csv", ".txt", ".md")):
            try:
                all_text.append(raw.decode("utf-8", errors="replace"))
            except Exception:
                pass
        elif name.endswith((".xlsx", ".xls", ".xlsm")):
            try:
                df = pd.read_excel(io.BytesIO(raw))
                all_text.append(df.to_string())
            except Exception:
                pass
    return "\n\n".join(all_text)[:80000]


def run_draw_request_analysis(client_type, client, project_info: dict, doc_text: str, extracted_data: dict = None) -> dict:
    """AI-analyse av trekkforespørsel mot budsjett og fremdrift."""
    # Build enhanced context from auto-extracted data
    extra_context = ""
    if extracted_data:
        extra_context = f"""

Auto-ekstrahert data fra dokumenter:
- Kalkyle: {json.dumps(extracted_data.get('kalkyle', {}), ensure_ascii=False, default=str)}
- Opptrekk: {json.dumps(extracted_data.get('opptrekk', {}), ensure_ascii=False, default=str)}
- Sjekkliste: {json.dumps(extracted_data.get('sjekkliste', {}), ensure_ascii=False, default=str)}
- Rapport: {json.dumps(extracted_data.get('rapport', {}), ensure_ascii=False, default=str)}
- Fakturaer: {json.dumps(extracted_data.get('fakturaer', []), ensure_ascii=False, default=str)}
- Identifiserte mangler: {json.dumps(extracted_data.get('mangler', []), ensure_ascii=False)}
- Antall bilag i opptrekk: {len(extracted_data.get('bilag', []))}
"""

    system_prompt = textwrap.dedent("""
    Du er en erfaren byggelånskontrollør i Norge. Du skal vurdere en trekkforespørsel
    fra en utbygger mot byggelånet. Du får prosjektinfo, dokumentgrunnlag og auto-ekstrahert
    strukturert data fra dokumentene.

    Bruk den auto-ekstraherte dataen til å gi presise tall. Verifiser at tallene stemmer
    overens mellom kalkyle, opptrekk og fakturaer.

    Returner KUN gyldig JSON med denne strukturen:
    {
        "sammendrag": "Kort oppsummering av trekkforespørselen (2-3 setninger)",
        "anbefalt_utbetaling_mnok": 0.0,
        "anbefalt_tilbakehold_mnok": 0.0,
        "godkjenningsstatus": "Anbefalt godkjent | Anbefalt med forbehold | Ikke anbefalt",
        "budsjett_vs_paloept": {
            "totalbudsjett_mnok": 0.0,
            "paloept_foer_trekk_mnok": 0.0,
            "dette_trekk_mnok": 0.0,
            "gjenstaaende_etter_trekk_mnok": 0.0,
            "forbruksprosent": 0.0
        },
        "fremdriftsvurdering": {
            "planlagt_fremdrift_pst": 0,
            "estimert_faktisk_fremdrift_pst": 0,
            "avvik_kommentar": "..."
        },
        "risikoer": [
            {"risiko": "...", "alvorlighet": "Lav|Middels|Høy|Kritisk", "tiltak": "..."}
        ],
        "kontrollpunkter": [
            {"punkt": "...", "status": "OK|Avvik|Mangler|Ikke vurdert", "kommentar": "..."}
        ],
        "dokumentasjonskontroll": [
            {"dokument": "...", "mottatt": true/false, "kommentar": "..."}
        ],
        "bilagskontroll": [
            {"leverandoer": "...", "beloep": 0.0, "kategori": "...", "verifisert": true/false, "kommentar": "..."}
        ],
        "vilkaar_for_utbetaling": ["..."],
        "anbefalinger": ["..."]
    }
    """)

    user_prompt = f"""
Prosjektinformasjon:
- Prosjekt: {project_info.get('navn', 'Ikke oppgitt')}
- Utbygger: {project_info.get('utbygger', 'Ikke oppgitt')}
- Totalbudsjett: {project_info.get('totalbudsjett_mnok', 0)} MNOK
- Byggelån innvilget: {project_info.get('byggelaan_mnok', 0)} MNOK
- Tidligere utbetalt: {project_info.get('tidligere_utbetalt_mnok', 0)} MNOK
- Forespurt trekk: {project_info.get('forespurt_trekk_mnok', 0)} MNOK
- Entrepriseform: {project_info.get('entrepriseform', 'Ikke oppgitt')}
- Planlagt ferdigstillelse: {project_info.get('ferdigstillelse', 'Ikke oppgitt')}
- Forhåndssalg/utleiegrad: {project_info.get('forhaandssalg_pst', 'Ikke oppgitt')}%
- Trekkforespørsel nr: {project_info.get('trekk_nr', 1)}
{extra_context}

Dokumentgrunnlag (utdrag):
{doc_text[:40000]}

Vurder trekkforespørselen og returner JSON.
"""

    try:
        if client_type == "openai":
            resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "system", "content": system_prompt}, {"role": "user", "content": user_prompt}],
                temperature=0.2, max_tokens=4000,
                response_format={"type": "json_object"},
            )
            return json.loads(resp.choices[0].message.content)
        elif client_type == "gemini":
            resp = client.generate_content(system_prompt + "\n\n" + user_prompt,
                                           generation_config={"temperature": 0.2, "max_output_tokens": 4000})
            text = resp.text.strip()
            text = re.sub(r"^```json\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
            return json.loads(text)
    except Exception as e:
        st.error(f"AI-analyse feilet: {e}")
    return {}


# ────────────────────────────────────────────────────────────────
# PDF REPORT
# ────────────────────────────────────────────────────────────────
class LoanControlPDF(FPDF if FPDF else object):
    """Corporate-grade PDF — McKinsey/BCG style for byggelånskontroll."""

    NAVY = (6, 17, 36); DARK_NAVY = (3, 10, 22); TEAL = (56, 194, 201)
    WARM = (245, 158, 11); GREEN = (34, 197, 94); RED = (239, 68, 68)
    WHITE = (255, 255, 255); LIGHT_GRAY = (245, 247, 251); MID_GRAY = (159, 176, 195)
    DARK_GRAY = (80, 100, 120); BODY_TEXT = (30, 40, 55); TABLE_HEAD = (12, 28, 50); TABLE_ALT = (240, 244, 250)

    def __init__(self):
        super().__init__("P", "mm", "A4")
        self.set_auto_page_break(auto=True, margin=28)
        self._has_unicode = self._add_fonts()
        self._logo_path = self._find_logo()
        self._logo_white_path = self._find_logo_white()
        self.accent = self.TEAL

    def _safe(self, text):
        if not text: return ""
        import unicodedata
        s = unicodedata.normalize('NFC', str(text))
        for old, new in [('\u2014','-'),('\u2013','-'),('\u2018',"'"),('\u2019',"'"),('\u201c','"'),('\u201d','"'),('\u2022','-')]:
            s = s.replace(old, new)
        return s if self._has_unicode else s.encode('latin-1', errors='replace').decode('latin-1')

    def _add_fonts(self):
        found = False
        for style, name in [("", "Inter-Regular.ttf"), ("B", "Inter-Bold.ttf")]:
            path = os.path.join(os.path.dirname(__file__), name)
            if os.path.exists(path):
                self.add_font("Inter", style, path, uni=True); found = True
        return found

    def _find_logo(self):
        for p in ["logo.png", os.path.join(os.path.dirname(__file__), "logo.png"), "/app/logo.png"]:
            if os.path.exists(p): return p
        return ""

    def _find_logo_white(self):
        for p in ["logo-white.png", os.path.join(os.path.dirname(__file__), "logo-white.png"), "/app/logo-white.png"]:
            if os.path.exists(p): return p
        return self._logo_path

    def _font(self, style="", size=10):
        try: self.set_font("Inter", style, size)
        except: self.set_font("Helvetica", style, size)

    def header(self):
        if self.page_no() <= 1: return
        y0 = 8
        if self._logo_path:
            try: self.image(self._logo_path, 10, y0, 22)
            except: self._font("B", 8); self.set_text_color(*self.TEAL); self.set_xy(10, y0+1); self.cell(22, 5, "BUILTLY")
        else:
            self._font("B", 8); self.set_text_color(*self.TEAL); self.set_xy(10, y0+1); self.cell(22, 5, "BUILTLY")
        self._font("B", 7); self.set_text_color(*self.MID_GRAY); self.set_xy(36, y0+1); self.cell(100, 5, self._safe("BYGGELÅNSKONTROLL"))
        self._font("", 7); self.set_text_color(*self.MID_GRAY); self.set_xy(150, y0+1); self.cell(50, 5, datetime.now().strftime("%d.%m.%Y"), align="R")
        self.set_draw_color(*self.TEAL); self.set_line_width(0.6); self.line(10, y0+7, 200, y0+7)
        self.set_draw_color(220, 225, 235); self.set_line_width(0.15); self.line(10, y0+7.8, 200, y0+7.8)
        self.set_y(y0 + 12)

    def footer(self):
        self.set_y(-18)
        self.set_draw_color(200, 210, 225); self.set_line_width(0.15); self.line(10, self.get_y(), 200, self.get_y())
        self.ln(2); self._font("", 6.5); self.set_text_color(*self.MID_GRAY)
        self.cell(120, 4, self._safe("KONFIDENSIELT - Kun for intern bruk"), align="L")
        pg = self.page_no() - 1
        if pg > 0: self.cell(0, 4, self._safe(f"Side {pg}"), align="R")
        self.ln(3); self._font("", 5.5); self.set_text_color(180, 190, 205)
        self.cell(0, 3, self._safe("Generert av Builtly | builtly.ai - Utkast, krever kvalitetssikring"), align="L")

    def cover_page(self, project_name, utbygger, trekk_nr):
        self.add_page()
        self.set_fill_color(*self.DARK_NAVY); self.rect(0, 0, 210, 297, style="F")
        # White logo on dark background
        cover_logo = self._logo_white_path or self._logo_path
        if cover_logo:
            try: self.image(cover_logo, 20, 25, 35)
            except: self._font("B", 14); self.set_text_color(*self.TEAL); self.set_xy(20, 25); self.cell(35, 10, "BUILTLY")
        else:
            self._font("B", 14); self.set_text_color(*self.TEAL); self.set_xy(20, 25); self.cell(35, 10, "BUILTLY")
        self.set_xy(20, 60); self.set_fill_color(*self.TEAL); self._font("B", 8); self.set_text_color(*self.DARK_NAVY)
        self.cell(42, 7, "  KONFIDENSIELT  ", fill=True, align="C")
        self.set_xy(20, 80); self._font("B", 32); self.set_text_color(*self.WHITE); self.cell(0, 14, self._safe("Byggelånskontroll"))
        self.set_xy(20, 98); self._font("", 14); self.set_text_color(*self.TEAL); self.cell(0, 8, self._safe(f"Trekkforespørsel #{trekk_nr}"))
        self.set_draw_color(*self.TEAL); self.set_line_width(1.2); self.line(20, 113, 90, 113)
        self.set_xy(20, 123); self._font("B", 20); self.set_text_color(*self.WHITE); self.multi_cell(170, 10, self._safe(project_name))
        y = self.get_y() + 8
        self.set_xy(20, y); self._font("", 11); self.set_text_color(*self.MID_GRAY); self.cell(0, 6, self._safe(f"Utbygger: {utbygger}"))
        box_y = 210
        self.set_fill_color(15, 30, 50); self.rect(20, box_y, 170, 40, style="F")
        self.set_draw_color(40, 60, 85); self.rect(20, box_y, 170, 40, style="D")
        items = [("Dato", datetime.now().strftime("%d.%m.%Y")), ("Klassifisering", "Konfidensielt"),
                 ("Utarbeidet av", "Builtly AI-assistert kontroll"), ("Status", "Utkast - krever faglig gjennomgang")]
        for i, (label, val) in enumerate(items):
            col_x = 25 + (i % 2) * 85; row_y = box_y + 6 + (i // 2) * 16
            self._font("B", 7); self.set_text_color(*self.TEAL); self.set_xy(col_x, row_y); self.cell(80, 4, label.upper())
            self._font("", 9); self.set_text_color(*self.WHITE); self.set_xy(col_x, row_y+5); self.cell(80, 4, self._safe(val))
        self.set_fill_color(*self.TEAL); self.rect(0, 293, 210, 4, style="F")

    def section_title(self, num, title):
        if self.get_y() > 225: self.add_page()
        self.ln(8)
        self._font("B", 8); self.set_fill_color(*self.TEAL); self.set_text_color(*self.WHITE)
        num_str = str(num); pill_w = max(8, len(num_str) * 3.5 + 5)
        self.cell(pill_w, 6, f" {num_str} ", fill=True, align="C"); self.cell(3, 6, "")
        self._font("B", 13); self.set_text_color(*self.NAVY)
        self.cell(0, 6, self._safe(title), new_x="LMARGIN", new_y="NEXT")
        self.set_draw_color(*self.TEAL); self.set_line_width(0.5); self.line(10, self.get_y()+1, 200, self.get_y()+1)
        self.ln(4)

    def body_text(self, text):
        if self.get_y() > 260: self.add_page()
        self._font("", 9.5); self.set_text_color(*self.BODY_TEXT)
        self.multi_cell(0, 5, self._safe(str(text))); self.ln(1.5)

    def key_value(self, key, value, highlight=False):
        if self.get_y() > 265: self.add_page()
        if highlight:
            self.set_fill_color(*self.TABLE_ALT); self.rect(10, self.get_y(), 190, 5.5, style="F")
        self._font("B", 8.5); self.set_text_color(*self.DARK_GRAY); self.cell(72, 5.5, self._safe(key))
        self._font("", 9.5); self.set_text_color(*self.NAVY); self.cell(0, 5.5, self._safe(str(value)), new_x="LMARGIN", new_y="NEXT")

    def status_box(self, status, text):
        if self.get_y() > 235: self.add_page()
        color_map = {"Anbefalt godkjent": self.GREEN, "Anbefalt med forbehold": self.WARM, "Ikke anbefalt": self.RED}
        color = color_map.get(status, self.TEAL)
        self.ln(3); y = self.get_y()
        self.set_fill_color(*color); self.rect(10, y, 190, 1.5, style="F")
        self.set_fill_color(min(color[0]+220,255), min(color[1]+220,255), min(color[2]+220,255))
        self.rect(10, y+1.5, 190, 22, style="F")
        self.set_draw_color(*color); self.set_line_width(0.3); self.rect(10, y, 190, 23.5, style="D")
        self._font("B", 13); self.set_text_color(*color); self.set_xy(16, y+4); self.cell(0, 7, self._safe(status))
        self._font("", 8.5); self.set_text_color(*self.BODY_TEXT); self.set_xy(16, y+12)
        self.multi_cell(176, 4.5, self._safe(text)); self.set_y(y + 27)

    def metric_row(self, metrics):
        if self.get_y() > 240: self.add_page()
        n = len(metrics)
        if n == 0: return
        card_w = (190 - (n-1)*4) / n; y = self.get_y()
        for i, (value, label, sublabel) in enumerate(metrics):
            x = 10 + i * (card_w + 4)
            self.set_fill_color(*self.LIGHT_GRAY); self.set_draw_color(220,225,235); self.set_line_width(0.2)
            self.rect(x, y, card_w, 22, style="DF")
            self.set_fill_color(*self.TEAL); self.rect(x, y, card_w, 1, style="F")
            self._font("B", 14); self.set_text_color(*self.NAVY); self.set_xy(x+4, y+3); self.cell(card_w-8, 7, self._safe(str(value)))
            self._font("B", 7); self.set_text_color(*self.DARK_GRAY); self.set_xy(x+4, y+11); self.cell(card_w-8, 4, self._safe(label.upper()))
            self._font("", 6.5); self.set_text_color(*self.MID_GRAY); self.set_xy(x+4, y+15.5); self.cell(card_w-8, 4, self._safe(sublabel))
        self.set_y(y + 26)

    def pro_table(self, headers, rows, col_widths=None):
        """Professional table with text wrapping and consistent alignment."""
        if self.get_y() > 240: self.add_page()
        n = len(headers)
        if col_widths is None: col_widths = [190/n]*n
        x_start = 10

        # Header row
        self.set_fill_color(*self.TABLE_HEAD); self.set_text_color(*self.WHITE); self._font("B", 7.5)
        for i, h in enumerate(headers):
            self.cell(col_widths[i], 7, self._safe(h), border=0, fill=True, align="L")
        self.ln()

        # Data rows with text wrapping
        self._font("", 8)
        row_h = 5.5
        for ri, row in enumerate(rows):
            if self.get_y() > 260: self.add_page()
            y_start = self.get_y()
            bg = self.TABLE_ALT if ri % 2 == 0 else self.WHITE

            # Calculate max height
            max_lines = 1
            for i, cell_val in enumerate(row):
                cs = self._safe(str(cell_val))
                chars_per_line = max(1, int(col_widths[i] * 3.2))
                lines = max(1, -(-len(cs) // chars_per_line))
                max_lines = max(max_lines, lines)
            cell_h = max_lines * row_h

            # Background
            self.set_fill_color(*bg)
            self.rect(x_start, y_start, 190, cell_h, style="F")

            # Draw cells with multi_cell for wrapping
            for i, cell_val in enumerate(row):
                cs = self._safe(str(cell_val))
                if any(k in cs.lower() for k in ["god","sterk","positiv","godkjent","mottatt","ja"]): self.set_text_color(*self.GREEN)
                elif any(k in cs.lower() for k in ["svak","negativ","ikke","kritisk","nei"]): self.set_text_color(*self.RED)
                elif any(k in cs.lower() for k in ["akseptabel","middels","forbehold"]): self.set_text_color(*self.WARM)
                else: self.set_text_color(*self.BODY_TEXT)

                x_pos = x_start + sum(col_widths[:i])
                self.set_xy(x_pos, y_start)
                self.multi_cell(col_widths[i], row_h, cs, border=0, align="L")

            self.set_y(y_start + cell_h)

        # Bottom line
        self.set_draw_color(200,210,225); self.set_line_width(0.3)
        self.line(x_start, self.get_y(), x_start + 190, self.get_y()); self.ln(2)

    def callout(self, title, text, tone="blue"):
        if self.get_y() > 250: self.add_page()
        tmap = {"blue":(self.TEAL,(230,248,250)), "green":(self.GREEN,(235,250,240)), "yellow":(self.WARM,(255,248,230)), "red":(self.RED,(255,235,235))}
        accent, bg = tmap.get(tone, tmap["blue"])
        y = self.get_y(); h = 16
        self.set_fill_color(*bg); self.set_draw_color(*accent); self.set_line_width(0.3); self.rect(10, y, 190, h, style="DF")
        self.set_fill_color(*accent); self.rect(10, y, 3, h, style="F")
        self._font("B", 8.5); self.set_text_color(*accent); self.set_xy(17, y+2); self.cell(0, 5, self._safe(title))
        self._font("", 8); self.set_text_color(*self.BODY_TEXT); self.set_xy(17, y+7.5); self.multi_cell(178, 4, self._safe(text))
        self.set_y(y + h + 3)

    def risk_table(self, risks):
        headers = ["Risiko", "Alvorlighet", "Tiltak"]
        rows = []
        for r in risks:
            rows.append([safe_get(r, "risiko", "-"), safe_get(r, "alvorlighet", "-"), safe_get(r, "tiltak", "-")])
        if rows:
            self.pro_table(headers, rows, [70, 25, 95])


def generate_loan_control_pdf(project_info, analysis) -> bytes:
    if not FPDF:
        return b""
    pdf = LoanControlPDF()
    pdf.alias_nb_pages()
    pdf.cover_page(
        project_info.get("navn", "Prosjekt"),
        project_info.get("utbygger", "-"),
        project_info.get("trekk_nr", 1),
    )
    pdf.add_page()

    # 1. Sammendrag og status
    pdf.section_title(1, "Sammendrag og anbefaling")
    status = safe_get(analysis, "godkjenningsstatus", "Ikke vurdert")
    pdf.status_box(status, safe_get(analysis, "sammendrag", ""))

    # 2. Budsjett vs påløpt — with metric cards
    pdf.section_title(2, "Budsjett vs. påløpt")
    bvp = safe_get(analysis, "budsjett_vs_paloept", {})
    if isinstance(bvp, dict):
        pdf.metric_row([
            (f"{safe_get(bvp, 'totalbudsjett_mnok', 0)} MNOK", "Totalbudsjett", "Samlet prosjektkost"),
            (f"{safe_get(bvp, 'dette_trekk_mnok', 0)} MNOK", "Forespurt trekk", "Dette trekk"),
            (f"{safe_get(bvp, 'forbruksprosent', 0)}%", "Forbruksprosent", "Andel forbrukt"),
            (f"{safe_get(bvp, 'gjenstaaende_etter_trekk_mnok', 0)} MNOK", "Gjenstående", "Etter dette trekk"),
        ])
        pdf.key_value("Påløpt før dette trekk:", f"{safe_get(bvp, 'paloept_foer_trekk_mnok', 0)} MNOK")
        pdf.key_value("Anbefalt tilbakehold:", f"{safe_get(bvp, 'anbefalt_tilbakehold_mnok', 0)} MNOK", highlight=True)

    # 3. Fremdrift
    pdf.section_title(3, "Fremdriftsvurdering")
    fv = safe_get(analysis, "fremdriftsvurdering", {})
    if isinstance(fv, dict):
        planlagt = safe_get(fv, 'planlagt_fremdrift_pst', 0)
        faktisk = safe_get(fv, 'estimert_faktisk_fremdrift_pst', 0)
        pdf.metric_row([
            (f"{planlagt}%", "Planlagt fremdrift", "Iht. tidsplan"),
            (f"{faktisk}%", "Faktisk fremdrift", "Estimert status"),
        ])
        avvik = safe_get(fv, "avvik_kommentar", "")
        if avvik:
            tone = "green" if faktisk >= planlagt else ("yellow" if faktisk >= planlagt * 0.85 else "red")
            pdf.callout("Fremdriftsavvik", avvik, tone)

    # 4. Kontrollpunkter — as professional table
    pdf.section_title(4, "Kontrollpunkter")
    kp_list = safe_get(analysis, "kontrollpunkter", [])
    if kp_list:
        headers = ["Kontrollpunkt", "Status", "Kommentar"]
        rows = []
        for kp in kp_list:
            if isinstance(kp, dict):
                rows.append([safe_get(kp, "punkt", ""), safe_get(kp, "status", "?"), safe_get(kp, "kommentar", "")])
        if rows:
            pdf.pro_table(headers, rows, [70, 25, 95])

    # 5. Risikoer
    pdf.section_title(5, "Risikovurdering")
    risks = safe_get(analysis, "risikoer", [])
    if risks:
        pdf.risk_table(risks)

    # 6. Dokumentasjonskontroll
    pdf.section_title(6, "Dokumentasjonskontroll")
    dok_list = safe_get(analysis, "dokumentasjonskontroll", [])
    if dok_list:
        headers = ["Dokument", "Status", "Kommentar"]
        rows = []
        for d in dok_list:
            if isinstance(d, dict):
                status_txt = "Mottatt" if safe_get(d, "mottatt", False) else "Ikke mottatt"
                rows.append([safe_get(d, "dokument", ""), status_txt, safe_get(d, "kommentar", "")])
        if rows:
            pdf.pro_table(headers, rows, [60, 30, 100])

    # 7. Vilkår og anbefalinger
    pdf.section_title(7, "Vilkår for utbetaling")
    vilkaar = safe_get(analysis, "vilkaar_for_utbetaling", [])
    if vilkaar:
        for i, v in enumerate(vilkaar, 1):
            pdf.body_text(f"{i}. {v}")

    pdf.section_title(8, "Anbefalinger")
    anbefalinger = safe_get(analysis, "anbefalinger", [])
    if anbefalinger:
        for i, a in enumerate(anbefalinger, 1):
            pdf.body_text(f"{i}. {a}")

    # Disclaimer
    pdf.ln(10)
    pdf.callout(
        "UTKAST - KREVER FAGLIG KONTROLL",
        "Analysen er automatisk generert av Builtly og skal gjennomgås av kvalifisert byggelånskontrollør før bruk.",
        "yellow"
    )

    return bytes(pdf.output())


# ────────────────────────────────────────────────────────────────
# PREMIUM CSS
# ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    #MainMenu, footer, header {visibility: hidden;}
    header[data-testid="stHeader"] {visibility: hidden; height: 0;}
    :root {
        --bg: #06111a; --panel: rgba(10,22,35,0.78); --stroke: rgba(120,145,170,0.18);
        --text: #f5f7fb; --muted: #9fb0c3; --soft: #c8d3df;
        --accent: #38bdf8; --accent-warm: #f59e0b; --radius-lg: 16px;
    }
    html, body, [class*="css"] { font-family: Inter, ui-sans-serif, system-ui, -apple-system, sans-serif; }
    .stApp { background-color: var(--bg) !important; color: var(--text); }
    .block-container { max-width: 1280px !important; padding-top: 1.5rem !important; }

    /* Hero */
    .hero-card { background: linear-gradient(135deg, rgba(10,22,35,0.95), rgba(16,32,50,0.95));
        border: 1px solid rgba(120,145,170,0.15); border-radius: 20px; padding: 2.5rem 2.8rem 2rem; margin-bottom: 2rem; }
    .hero-eyebrow { font-size: 0.78rem; font-weight: 700; letter-spacing: 0.12em; text-transform: uppercase; color: #38bdf8; margin-bottom: 0.4rem; }
    .hero-title { font-size: 1.85rem; font-weight: 800; line-height: 1.2; color: #f5f7fb; margin-bottom: 0.75rem; }
    .hero-subtitle { font-size: 0.98rem; color: #9fb0c3; line-height: 1.6; max-width: 750px; }
    .hero-pills { display: flex; gap: 8px; flex-wrap: wrap; margin-top: 1.1rem; }
    .hero-pill { background: rgba(56,194,201,0.08); border: 1px solid rgba(56,194,201,0.25); border-radius: 20px; padding: 4px 14px; font-size: 0.8rem; font-weight: 600; color: #38bdf8; }
    .hero-badge { display: inline-block; background: rgba(245,158,11,0.12); border: 1px solid rgba(245,158,11,0.35); border-radius: 6px; padding: 2px 10px; font-size: 0.72rem; font-weight: 700; color: #f59e0b; text-transform: uppercase; letter-spacing: 0.08em; margin-top: 1rem; }

    /* Metrics */
    .metric-row { display: flex; gap: 16px; margin-bottom: 1.5rem; flex-wrap: wrap; }
    .metric-card { flex: 1; min-width: 200px; background: rgba(10,22,35,0.6); border: 1px solid rgba(120,145,170,0.18); border-radius: 14px; padding: 1.1rem 1.3rem; }
    .metric-card .mc-value { font-size: 1.6rem; font-weight: 800; color: #38bdf8; margin-bottom: 2px; }
    .metric-card .mc-label { font-size: 0.82rem; font-weight: 700; color: #c8d3df; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px; }
    .metric-card .mc-desc { font-size: 0.78rem; color: #9fb0c3; line-height: 1.4; }

    /* Sections */
    .section-header { margin-top: 2.5rem; margin-bottom: 1rem; }
    .section-header h3 { color: #f5f7fb !important; font-weight: 750 !important; font-size: 1.2rem !important; margin-bottom: 4px !important; }
    .section-header p { color: #9fb0c3 !important; font-size: 0.9rem !important; }
    .section-badge { display: inline-block; background: rgba(56,194,201,0.1); border: 1px solid rgba(56,194,201,0.25); border-radius: 6px; padding: 1px 8px; font-size: 0.7rem; font-weight: 700; color: #38bdf8; text-transform: uppercase; letter-spacing: 0.06em; margin-bottom: 4px; }

    /* Auto-extracted indicator */
    .auto-badge { display: inline-block; background: rgba(34,197,94,0.1); border: 1px solid rgba(34,197,94,0.3); border-radius: 4px;
        padding: 1px 6px; font-size: 0.65rem; font-weight: 700; color: #22c55e; text-transform: uppercase; letter-spacing: 0.05em; margin-left: 6px; }
    .manual-badge { display: inline-block; background: rgba(245,158,11,0.1); border: 1px solid rgba(245,158,11,0.3); border-radius: 4px;
        padding: 1px 6px; font-size: 0.65rem; font-weight: 700; color: #f59e0b; text-transform: uppercase; letter-spacing: 0.05em; margin-left: 6px; }
    .doc-card { background: rgba(10,22,35,0.6); border: 1px solid rgba(120,145,170,0.18); border-radius: 10px;
        padding: 0.7rem 1rem; margin-bottom: 0.5rem; display: flex; justify-content: space-between; align-items: center; }
    .doc-name { color: #f5f7fb; font-weight: 600; font-size: 0.88rem; }
    .doc-type { color: #38bdf8; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; }
    .doc-fields { color: #9fb0c3; font-size: 0.75rem; }
    .extraction-summary { background: rgba(34,197,94,0.06); border: 1px solid rgba(34,197,94,0.2); border-radius: 14px;
        padding: 1.2rem 1.5rem; margin: 1rem 0; }
    .extraction-summary h4 { color: #22c55e !important; margin-bottom: 0.5rem !important; font-size: 1rem !important; }
    .extraction-summary .stat { color: #c8d3df; font-size: 0.88rem; line-height: 1.6; }

    /* Inputs */
    .stTextInput input, .stTextArea textarea, .stNumberInput input, .stSelectbox > div > div,
    .stMultiSelect > div > div { background-color: rgba(10,22,35,0.6) !important; color: #f5f7fb !important;
        border: 1px solid rgba(120,145,170,0.2) !important; border-radius: 12px !important; }
    div[data-baseweb='base-input'], div[data-baseweb='input'] {
        background-color: rgba(10,22,35,0.6) !important; border-color: rgba(120,145,170,0.2) !important; }
    div[data-baseweb='base-input'] input, div[data-baseweb='input'] input,
    .stTextInput input, .stTextArea textarea, .stNumberInput input {
        color: #f5f7fb !important; -webkit-text-fill-color: #f5f7fb !important;
        caret-color: #38c2c9 !important; background: transparent !important; }
    input:-webkit-autofill, textarea:-webkit-autofill {
        -webkit-box-shadow: 0 0 0px 1000px rgba(10,22,35,0.95) inset !important;
        -webkit-text-fill-color: #f5f7fb !important; caret-color: #38c2c9 !important; }
    input::placeholder, textarea::placeholder {
        color: rgba(159,176,195,0.6) !important; -webkit-text-fill-color: rgba(159,176,195,0.6) !important; }
    .stNumberInput div[data-baseweb='base-input'] { background-color: rgba(10,22,35,0.6) !important; }
    ul[data-baseweb='menu'], ul[data-baseweb='menu'] li {
        background-color: rgba(10,22,35,0.95) !important; color: #f5f7fb !important; }
    ul[data-baseweb='menu'] li:hover, li[aria-selected='true'] { background-color: rgba(56,194,201,0.12) !important; }

    .stSelectbox label, .stMultiSelect label, .stTextInput label, .stTextArea label,
    .stNumberInput label, .stFileUploader label, .stToggle label, .stRadio label,
    .stDateInput label { color: #c8d3df !important; font-weight: 600 !important; }
    div[data-baseweb="select"] > div { background-color: rgba(10,22,35,0.6) !important; border-color: rgba(120,145,170,0.2) !important; }

    /* Tabs */
    .stTabs [data-baseweb="tab-list"] { gap: 6px; border-bottom: 1px solid rgba(120,145,170,0.15); }
    .stTabs [data-baseweb="tab"] { background: transparent !important; color: #9fb0c3 !important; border-radius: 10px 10px 0 0 !important; padding: 8px 18px !important; font-weight: 600 !important; }
    .stTabs [aria-selected="true"] { background: rgba(56,194,201,0.08) !important; color: #38bdf8 !important; border-bottom: 2px solid #38bdf8 !important; }

    /* Buttons */
    button[kind="primary"], .stFormSubmitButton > button {
        background: linear-gradient(135deg, rgba(56,194,201,0.96), rgba(120,220,225,0.96)) !important;
        color: #041018 !important; border: none !important; font-weight: 750 !important;
        border-radius: 12px !important; padding: 12px 24px !important; font-size: 1.05rem !important; }
    button[kind="secondary"] { background-color: rgba(255,255,255,0.04) !important; color: #c8d3df !important;
        border: 1px solid rgba(120,145,170,0.25) !important; border-radius: 10px !important; font-weight: 600 !important; }
    .stDownloadButton > button { background-color: rgba(255,255,255,0.04) !important; color: #c8d3df !important;
        border: 1px solid rgba(120,145,170,0.25) !important; border-radius: 10px !important; font-weight: 600 !important; }

    /* DataFrame */
    .stDataFrame { border-radius: 12px; overflow: hidden; }
    .stDataFrame [data-testid="stDataFrameResizable"] { border: 1px solid rgba(120,145,170,0.15) !important; border-radius: 12px !important; }
    .stDataFrame [data-testid="glideDataEditor"], .dvn-scroller, .dvn-scroller div { background-color: #0c1c2c !important; }
    .stDataFrame th, .stDataFrame [role="columnheader"] { background-color: #112236 !important; color: #c8d3df !important; }
    .stDataFrame td, .stDataFrame [role="gridcell"] { background-color: #0c1c2c !important; color: #f5f7fb !important; }

    /* Alerts */
    .stAlert, div[data-testid="stAlert"], div[role="alert"] { background-color: #112236 !important; color: #f5f7fb !important; }
    .stAlert p, .stAlert span, div[role="alert"] p, div[role="alert"] span { color: #f5f7fb !important; }

    /* File uploader */
    [data-testid="stFileUploaderDropzone"] { background-color: #0c1c2c !important; border-color: rgba(120,145,170,0.18) !important; color: #f5f7fb !important; }
    [data-testid="stFileUploaderFile"] { background-color: #112236 !important; color: #f5f7fb !important; }
    [data-testid="stFileUploaderFile"] span { color: #f5f7fb !important; }

    /* Date picker */
    div[data-baseweb="calendar"], div[data-baseweb="calendar"] * { background-color: #162a42 !important; color: #f5f7fb !important; }
    div[data-baseweb="datepicker"] { background-color: #162a42 !important; }

    /* Tooltips/popups */
    div[data-baseweb="tooltip"] > div, div[data-baseweb="popover"] > div { background-color: #162a42 !important; color: #f5f7fb !important; }

    /* Expander */
    details[data-testid="stExpander"], details[data-testid="stExpander"] summary,
    details[data-testid="stExpander"] > div { background-color: #0c1c2c !important; color: #f5f7fb !important; }

    /* Number input stepper */
    .stNumberInput button { background-color: #112236 !important; color: #c8d3df !important; }

    /* Toggle/checkbox/radio */
    .stToggle span, .stCheckbox span, .stRadio span { color: #f5f7fb !important; }

    /* Scrollbars */
    ::-webkit-scrollbar { width: 8px; height: 8px; }
    ::-webkit-scrollbar-track { background: #0c1c2c; }
    ::-webkit-scrollbar-thumb { background: rgba(120,145,170,0.3); border-radius: 4px; }

    /* Catch-all white backgrounds */
    .stApp div[style*="background-color: white"], .stApp div[style*="background: rgb(255"] { background-color: #0c1c2c !important; }

    /* Disclaimer */
    .disclaimer-banner { background: rgba(245,158,11,0.06); border: 1px solid rgba(245,158,11,0.25); border-radius: 14px; padding: 1.1rem 1.4rem; margin-top: 2rem; }
    .disclaimer-banner .db-title { font-weight: 700; font-size: 0.9rem; color: #f59e0b; margin-bottom: 4px; }
    .disclaimer-banner .db-text { font-size: 0.82rem; color: #9fb0c3; line-height: 1.5; }

    /* Markdown */
    .stMarkdown, .stMarkdown p, .stMarkdown li, .stMarkdown h1, .stMarkdown h2, .stMarkdown h3, .stMarkdown h4 { color: #f5f7fb !important; }

    /* Status colors */
    .status-green { color: #22c55e; font-weight: 700; }
    .status-yellow { color: #f59e0b; font-weight: 700; }
    .status-red { color: #ef4444; font-weight: 700; }
</style>
""", unsafe_allow_html=True)


# ────────────────────────────────────────────────────────────────
# BACK BUTTON + LOGO
# ────────────────────────────────────────────────────────────────
top_l, top_r = st.columns([6, 1])
with top_l:
    logo = logo_data_uri()
    if logo:
        render_html(f'<img src="{logo}" class="brand-logo" alt="Builtly">')
with top_r:
    if st.button("← Tilbake", type="secondary", key="back_btn"):
        try:
            st.switch_page("pages/Project.py")
        except Exception:
            st.info("Naviger tilbake til prosjektoversikten manuelt.")


# ────────────────────────────────────────────────────────────────
# HERO
# ────────────────────────────────────────────────────────────────
render_hero(
    eyebrow="Byggelånskontroll v3",
    title="Last opp dokumentasjon — vi leser og fyller ut automatisk.",
    subtitle=(
        "Last opp trekkanmodning, kalkyle, sjekkliste, byggelånsrapport og fakturaer. "
        "Modulen leser dokumentene, ekstraherer nøkkeltall og fyller ut alle felt automatisk. "
        "Du trenger bare å verifisere og supplere det som eventuelt mangler."
    ),
    pills=["Auto-ekstraksjon", "Dokumentklassifisering", "Smart pre-utfylling", "Bilagskontroll", "PDF-rapport"],
    badge="Oppgradert med auto-lesing",
)


# ────────────────────────────────────────────────────────────────
# MAIN LAYOUT
# ────────────────────────────────────────────────────────────────

# STEP 1: Upload documents FIRST
render_section("1. Last opp dokumentasjon", "Last opp alle tilgjengelige dokumenter. Modulen klassifiserer og leser dem automatisk.", "Dokumenter")

uploads = st.file_uploader(
    "Trekkanmodning (xlsx), prosjektkalkyle, sjekkliste, byggelånsrapport, fakturaer, byggeplassbilder",
    type=["pdf", "xlsx", "xls", "xlsm", "csv", "docx", "jpg", "jpeg", "png", "zip"],
    accept_multiple_files=True,
    key="loan_uploads",
)

# Auto-extract when files are uploaded
extracted_data = None
if uploads and "auto_extracted" not in st.session_state:
    with st.spinner("Leser og klassifiserer dokumenter..."):
        extracted_data = auto_extract_all(uploads)
        st.session_state["auto_extracted"] = extracted_data

if "auto_extracted" in st.session_state:
    extracted_data = st.session_state["auto_extracted"]

# Force re-extraction if files change
if uploads and extracted_data:
    current_files = sorted([f.name for f in uploads])
    prev_files = sorted([d['filename'] for d in extracted_data.get('documents', [])])
    if current_files != prev_files:
        with st.spinner("Leser og klassifiserer dokumenter..."):
            extracted_data = auto_extract_all(uploads)
            st.session_state["auto_extracted"] = extracted_data

# Show extraction results
if extracted_data and extracted_data.get('documents'):
    docs = extracted_data['documents']
    proj = extracted_data.get('project', {})
    confidence = extracted_data.get('confidence', {})

    # Document classification cards
    type_labels = {
        'kalkyle': '📊 Kalkyle',
        'sjekkliste': '✅ Sjekkliste',
        'byggelaansrapport': '📋 Byggelånsrapport',
        'opptrekk': '💰 Opptrekk',
        'faktura': '🧾 Faktura',
        'epost': '📧 E-post',
        'kontrakt': '📄 Kontrakt',
        'fremdrift': '📈 Fremdrift',
        'betalingsplan': '💳 Betalingsplan',
        'bilde': '📷 Bilde',
        'ukjent': '❓ Ukjent',
        'excel': '📊 Excel',
        'tekst': '📝 Tekst',
    }

    docs_html = ""
    for doc in docs:
        type_label = type_labels.get(doc['type'], f"📄 {doc['type']}")
        fields_count = len(doc.get('extracted_fields', []))
        fields_text = f"{fields_count} felt ekstrahert" if fields_count > 0 else "Ingen strukturerte data"
        docs_html += f"""<div class="doc-card">
            <div><span class="doc-name">{doc['filename']}</span></div>
            <div><span class="doc-type">{type_label}</span>
            <span class="doc-fields"> — {fields_text}</span></div>
        </div>"""
    render_html(docs_html)

    # Extraction summary
    auto_filled = len(confidence)
    total_fields = 15  # Approximate total fields
    render_html(f"""
    <div class="extraction-summary">
        <h4>✅ Auto-ekstraksjon fullført</h4>
        <div class="stat">{len(docs)} dokumenter klassifisert • {auto_filled} av ~{total_fields} felt fylt automatisk</div>
        <div class="stat" style="margin-top:4px;color:#9fb0c3;">Felt med <span class="auto-badge">AUTO</span> er hentet fra dokumentene.
        Felt med <span class="manual-badge">MANUELL</span> må fylles inn.</div>
    </div>""")

    if extracted_data.get('mangler'):
        mangler_html = "".join(f"<li style='color:#ef4444;font-size:0.85rem;'>{m}</li>" for m in extracted_data['mangler'])
        render_html(f"""
        <div style="background:rgba(239,68,68,0.06);border:1px solid rgba(239,68,68,0.2);border-radius:12px;padding:1rem 1.3rem;margin:0.5rem 0;">
            <div style="font-weight:700;color:#ef4444;font-size:0.9rem;margin-bottom:6px;">⚠️ Manglende dokumentasjon identifisert</div>
            <ul style="margin:0;padding-left:1.2rem;">{mangler_html}</ul>
        </div>""")


# STEP 2: Pre-filled form (with auto-extracted values)
render_section("2. Verifiser prosjektdata", "Feltene under er forhåndsutfylt basert på dokumentene. Korriger eller fyll inn manglende data.", "Verifiser")

proj = extracted_data.get('project', {}) if extracted_data else {}
confidence = extracted_data.get('confidence', {}) if extracted_data else {}

def field_label(label: str, field_key: str) -> str:
    """Returns label with auto/manual badge."""
    if field_key in confidence:
        return f"{label}"
    return f"{label}"

col1, col2 = st.columns(2)
with col1:
    default_name = proj.get('prosjektnavn', '')
    prosjekt_navn = st.text_input(
        field_label("Prosjektnavn", "prosjektnavn"),
        value=default_name,
        placeholder="Prosjektnavn",
        help="🟢 Auto-fylt" if 'prosjektnavn' in confidence else "🟡 Fyll inn manuelt"
    )

    default_utbygger = proj.get('utbygger', '')
    utbygger = st.text_input(
        field_label("Utbygger / låntaker", "utbygger"),
        value=default_utbygger,
        placeholder="Selskap AS",
        help="🟢 Auto-fylt" if 'utbygger' in confidence else "🟡 Fyll inn manuelt"
    )

    default_budget = proj.get('totalbudsjett', 0) / 1_000_000 if proj.get('totalbudsjett', 0) > 100_000 else proj.get('totalbudsjett', 0)
    totalbudsjett = st.number_input(
        "Totalbudsjett (MNOK)",
        min_value=0.0,
        value=float(round(default_budget, 1)),
        step=1.0, format="%.1f",
        help="🟢 Auto-fylt fra kalkyle" if 'totalbudsjett' in confidence else "🟡 Fyll inn manuelt"
    )

    default_loan = proj.get('byggelaan', 0) / 1_000_000 if proj.get('byggelaan', 0) > 100_000 else proj.get('byggelaan', 0)
    byggelaan = st.number_input(
        "Byggelån innvilget (MNOK)",
        min_value=0.0,
        value=float(round(default_loan, 1)),
        step=1.0, format="%.1f",
        help="🟢 Auto-fylt fra opptrekk" if 'byggelaan' in confidence else "🟡 Fyll inn manuelt"
    )

with col2:
    default_prev = proj.get('tidligere_utbetalt', 0) / 1_000_000 if proj.get('tidligere_utbetalt', 0) > 100_000 else proj.get('tidligere_utbetalt', 0)
    tidligere_utbetalt = st.number_input(
        "Tidligere utbetalt (MNOK)",
        min_value=0.0,
        value=float(round(default_prev, 1)),
        step=1.0, format="%.1f",
        help="🟢 Auto-fylt fra opptrekk" if 'tidligere_utbetalt' in confidence else "🟡 Fyll inn manuelt"
    )

    default_trekk = proj.get('forespurt_trekk', 0) / 1_000_000 if proj.get('forespurt_trekk', 0) > 100_000 else proj.get('forespurt_trekk', 0)
    forespurt_trekk = st.number_input(
        "Forespurt trekk dette (MNOK)",
        min_value=0.0,
        value=float(round(default_trekk, 2)),
        step=0.5, format="%.2f",
        help="🟢 Auto-fylt fra opptrekk" if 'forespurt_trekk' in confidence else "🟡 Fyll inn manuelt"
    )

    default_entreprise = 'Totalentreprise' if proj.get('entrepriseform') == 'Totalentreprise' else 'Totalentreprise'
    entreprise_options = ["Totalentreprise", "Hovedentreprise", "Delte entrepriser", "Byggherrestyrt", "Annet"]
    entrepriseform = st.selectbox(
        "Entrepriseform",
        entreprise_options,
        index=entreprise_options.index(default_entreprise) if default_entreprise in entreprise_options else 0,
        help="🟢 Auto-fylt fra sjekkliste" if 'entrepriseform' in confidence else "🟡 Velg manuelt"
    )

    default_trekk_nr = proj.get('trekk_nr', 1)
    trekk_nr = st.number_input(
        "Trekkforespørsel nr.",
        min_value=1,
        value=int(default_trekk_nr),
        step=1,
        help="🟢 Auto-fylt fra opptrekk" if 'trekk_nr' in confidence else "🟡 Fyll inn manuelt"
    )

c3, c4 = st.columns(2)
with c3:
    ferdigstillelse = st.date_input("Planlagt ferdigstillelse", value=date(2027, 6, 30))
    default_salg = proj.get('salgsgrad', 0)
    forhaandssalg = st.number_input(
        "Forhåndssalg/utleiegrad (%)",
        min_value=0, max_value=100,
        value=int(default_salg),
        step=5,
        help="🟢 Auto-fylt fra kalkyle" if 'salgsgrad' in confidence else "🟡 Fyll inn manuelt"
    )
with c4:
    egenkapital_pst = st.number_input("Egenkapitalandel (%)", min_value=0, max_value=100, value=25, step=1)
    garanti_type = st.selectbox("Garantistillelse", ["Bankgaranti §12", "Selvskyldnergaranti", "Eiendomspant", "Kombinert", "Annet"])


# STEP 3: Show extracted line items if available
if extracted_data and extracted_data.get('bilag'):
    render_section("3. Bilagsoversikt", "Bilag ekstrahert fra opptrekk-filen. Verifiser at beløpene stemmer.", "Bilag")

    bilag_df = pd.DataFrame(extracted_data['bilag'])
    # Clean up display
    display_cols = ['leverandoer', 'beloep', 'kategori', 'dato', 'fakturanr', 'dokument_id']
    available_cols = [c for c in display_cols if c in bilag_df.columns]
    if available_cols:
        display_df = bilag_df[available_cols].copy()
        col_rename = {'leverandoer': 'Leverandør', 'beloep': 'Beløp', 'kategori': 'Kategori',
                     'dato': 'Dato', 'fakturanr': 'Fakturanr', 'dokument_id': 'DokumentID'}
        display_df = display_df.rename(columns={k: v for k, v in col_rename.items() if k in display_df.columns})
        if 'Beløp' in display_df.columns:
            total = display_df['Beløp'].sum()
            st.markdown(f"**Sum bilag: {fmt_nok(total)}**")
        st.dataframe(display_df, use_container_width=True, hide_index=True)

# STEP 4: Byggeplassbilder
render_section("4. Byggeplassbilder", "Last opp bilder fra byggeplassen som dokumenterer fremdrift.", "Bilder")

render_html('''<div style="background:rgba(56,194,201,0.06);border:1px solid rgba(56,194,201,0.18);border-radius:12px;padding:0.7rem 1.1rem;margin-bottom:0.8rem;">
    <div style="font-size:0.82rem;color:#9fb0c3;">Byggeplassbilder underbygger fremdriftsvurderingen og inkluderes i rapporten.
    Bilder kan også lastes opp sammen med øvrig dokumentasjon i steg 1.</div>
</div>''')

# Collect images from uploads + dedicated uploader
byggeplass_bilder = []
if uploads:
    for f in uploads:
        if f.name.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
            byggeplass_bilder.append(f)

ekstra_bilder = st.file_uploader(
    "Last opp byggeplassbilder",
    type=["jpg", "jpeg", "png", "gif", "webp"],
    accept_multiple_files=True,
    key="byggeplass_bilder",
    label_visibility="collapsed",
)
if ekstra_bilder:
    byggeplass_bilder.extend(ekstra_bilder)

if byggeplass_bilder:
    render_html(f'<div style="font-size:0.82rem;color:#38bdf8;font-weight:600;margin-bottom:0.5rem;">{len(byggeplass_bilder)} bilde(r) lastet opp</div>')
    # Display in grid
    img_cols = st.columns(min(len(byggeplass_bilder), 3))
    for idx, img_file in enumerate(byggeplass_bilder):
        with img_cols[idx % 3]:
            st.image(img_file, caption=img_file.name, use_container_width=True)

st.session_state["_byggeplass_bilder"] = byggeplass_bilder

# STEP 5: Additional context
render_section("5. Tilleggsinformasjon", "Eventuelle spesielle forhold byggelånskontrollør bør være oppmerksom på.", "Kontekst")

fokus = st.text_area(
    "Spesielle forhold å fokusere på",
    value="",
    placeholder="F.eks. entreprenør har varslet tillegg, forsinkelser, endringer i entreprisekontrakt, etc.",
    height=90,
)

run_analysis = st.button("Analyser trekkforespørsel", type="primary", use_container_width=True)


# ────────────────────────────────────────────────────────────────
# ANALYSIS
# ────────────────────────────────────────────────────────────────
if run_analysis:
    project_info = {
        "navn": prosjekt_navn or "Ikke oppgitt",
        "utbygger": utbygger or "Ikke oppgitt",
        "totalbudsjett_mnok": totalbudsjett,
        "byggelaan_mnok": byggelaan,
        "tidligere_utbetalt_mnok": tidligere_utbetalt,
        "forespurt_trekk_mnok": forespurt_trekk,
        "entrepriseform": entrepriseform,
        "ferdigstillelse": str(ferdigstillelse),
        "forhaandssalg_pst": forhaandssalg,
        "egenkapital_pst": egenkapital_pst,
        "garanti_type": garanti_type,
        "trekk_nr": trekk_nr,
        "fokus": fokus,
    }

    client_type, client = get_ai_client()
    if not client:
        st.error("Ingen AI-nøkkel konfigurert. Sett OPENAI_API_KEY eller GOOGLE_API_KEY i miljøvariablene.")
        st.stop()

    doc_text = ""
    if uploads:
        with st.spinner("Leser dokumenter..."):
            doc_text = extract_text_from_uploads(uploads)

    with st.spinner("Analyserer trekkforespørsel..."):
        analysis = run_draw_request_analysis(client_type, client, project_info, doc_text, extracted_data)

    if not analysis:
        st.error("Analysen returnerte ingen resultater. Sjekk dokumentgrunnlaget og prøv igjen.")
        st.stop()

    st.session_state["loan_analysis"] = analysis
    st.session_state["loan_project_info"] = project_info

# ── Display results ──
if "loan_analysis" in st.session_state:
    analysis = st.session_state["loan_analysis"]
    project_info = st.session_state.get("loan_project_info", {})

    render_section("Resultat", "Trekkanbefaling basert på innsendt dokumentasjon og prosjektdata.", "Analyse")

    # Status banner
    status = safe_get(analysis, "godkjenningsstatus", "Ikke vurdert")
    status_class = {"Anbefalt godkjent": "status-green", "Anbefalt med forbehold": "status-yellow", "Ikke anbefalt": "status-red"}.get(status, "")
    render_html(f"""
    <div style="background:rgba(10,22,35,0.7);border:1px solid rgba(120,145,170,0.2);border-radius:16px;padding:1.5rem 2rem;margin-bottom:1.5rem;">
        <div style="font-size:0.78rem;color:#9fb0c3;text-transform:uppercase;font-weight:700;letter-spacing:0.08em;margin-bottom:4px;">Anbefaling</div>
        <div class="{status_class}" style="font-size:1.5rem;margin-bottom:6px;">{status}</div>
        <div style="color:#c8d3df;font-size:0.92rem;line-height:1.6;">{safe_get(analysis, 'sammendrag', '')}</div>
    </div>""")

    # Budget metrics
    bvp = safe_get(analysis, "budsjett_vs_paloept", {})
    if isinstance(bvp, dict):
        render_metric_cards([
            (f"{safe_get(bvp, 'totalbudsjett_mnok', 0)} MNOK", "Totalbudsjett", "Samlet prosjektbudsjett"),
            (f"{safe_get(bvp, 'paloept_foer_trekk_mnok', 0)} MNOK", "Påløpt før trekk", "Akkumulert forbruk"),
            (f"{safe_get(bvp, 'dette_trekk_mnok', 0)} MNOK", "Dette trekket", "Forespurt beløp"),
            (f"{safe_get(bvp, 'forbruksprosent', 0)}%", "Forbruksprosent", "Andel av totalbudsjett brukt"),
        ])

    # Tabs
    tabs = st.tabs(["Kontrollpunkter", "Risikoer", "Dokumentkontroll", "Fremdrift", "Bilagskontroll", "Vilkår", "Eksport"])

    with tabs[0]:
        kp_list = safe_get(analysis, "kontrollpunkter", [])
        if kp_list:
            rows = []
            for kp in kp_list:
                if isinstance(kp, dict):
                    rows.append({"Kontrollpunkt": safe_get(kp, "punkt"), "Status": safe_get(kp, "status"), "Kommentar": safe_get(kp, "kommentar")})
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with tabs[1]:
        risks = safe_get(analysis, "risikoer", [])
        if risks:
            rows = []
            for r in risks:
                if isinstance(r, dict):
                    rows.append({"Risiko": safe_get(r, "risiko"), "Alvorlighet": safe_get(r, "alvorlighet"), "Tiltak": safe_get(r, "tiltak")})
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with tabs[2]:
        docs = safe_get(analysis, "dokumentasjonskontroll", [])
        if docs:
            rows = []
            for d in docs:
                if isinstance(d, dict):
                    rows.append({"Dokument": safe_get(d, "dokument"), "Mottatt": "✓" if safe_get(d, "mottatt", False) else "✗", "Kommentar": safe_get(d, "kommentar")})
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with tabs[3]:
        fv = safe_get(analysis, "fremdriftsvurdering", {})
        if isinstance(fv, dict):
            fc1, fc2 = st.columns(2)
            with fc1:
                render_metric_cards([
                    (f"{safe_get(fv, 'planlagt_fremdrift_pst', 0)}%", "Planlagt fremdrift", "Iht. fremdriftsplan"),
                ])
            with fc2:
                render_metric_cards([
                    (f"{safe_get(fv, 'estimert_faktisk_fremdrift_pst', 0)}%", "Estimert faktisk", "Basert på dokumentasjon"),
                ])
            st.markdown(safe_get(fv, "avvik_kommentar", ""))

    with tabs[4]:
        # Bilagskontroll tab (new)
        bk_list = safe_get(analysis, "bilagskontroll", [])
        if bk_list:
            rows = []
            for bk in bk_list:
                if isinstance(bk, dict):
                    rows.append({
                        "Leverandør": safe_get(bk, "leverandoer"),
                        "Beløp": safe_get(bk, "beloep"),
                        "Kategori": safe_get(bk, "kategori"),
                        "Verifisert": "✓" if safe_get(bk, "verifisert", False) else "✗",
                        "Kommentar": safe_get(bk, "kommentar"),
                    })
            if rows:
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        else:
            st.info("Ingen bilagskontroll tilgjengelig. Last opp fakturaer og opptrekk for detaljert kontroll.")

    with tabs[5]:
        vilkaar = safe_get(analysis, "vilkaar_for_utbetaling", [])
        if vilkaar:
            for i, v in enumerate(vilkaar, 1):
                st.markdown(f"**{i}.** {v}")
        st.markdown("---")
        st.markdown("**Anbefalinger:**")
        for a in safe_get(analysis, "anbefalinger", []):
            st.markdown(f"• {a}")

    with tabs[6]:
        # PDF download
        pdf_bytes = generate_loan_control_pdf(project_info, analysis)
        if pdf_bytes:
            st.download_button(
                "Last ned PDF-rapport",
                data=pdf_bytes,
                file_name=f"byggelanskontroll_{project_info.get('navn', 'prosjekt').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
        # JSON
        st.download_button(
            "Last ned analyse (JSON)",
            data=json.dumps({"prosjekt": project_info, "analyse": analysis}, ensure_ascii=False, indent=2),
            file_name=f"byggelanskontroll_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json",
            use_container_width=True,
        )


# ────────────────────────────────────────────────────────────────
# DISCLAIMER
# ────────────────────────────────────────────────────────────────
render_html("""
<div class="disclaimer-banner" style="margin-top: 2rem;">
    <div class="db-title">Utkast — krever faglig kontroll</div>
    <div class="db-text">
        Analysen er automatisk generert basert på innsendt dokumentasjon og oppgitte prosjektdata.
        Resultatet skal gjennomgås og verifiseres av kvalifisert byggelånskontrollør før det benyttes
        som grunnlag for utbetalingsbeslutning.
    </div>
</div>
""")
